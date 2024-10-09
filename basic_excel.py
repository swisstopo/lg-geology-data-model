import sys
import textwrap

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from ruamel.yaml import YAML


def split_long_string(text, max_length=75):
    if text is None:
        return ""
    # Use textwrap to split the text at max_length, keeping words intact
    # text = text.replace("\n", "@@")
    wrapped_lines = textwrap.wrap(text, width=max_length)

    # Join the lines, replacing single line feeds with double line feeds
    result = "\n".join(wrapped_lines)

    # Replace existing single '\n' in the original text with '\n\n'
    result = result.replace("@@", "\n\n")

    return result


yaml = YAML()

# Load data from the YAML file
with open("datamodel.yaml", "r") as file:
    yaml_data = yaml.load(file)

# Prepare the data for DataFrame
data = []
previous_theme = "dummy"
previous_class = "dummy"
previous_abrev = "dummy"
previous_table = "dummy"
previous_description_de = "dummy"
previous_description_fr = "dummy"

# Create an Excel file
file_path = "basic_excel.xlsx"
with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
    current_row = 1  # Starting row for writing data
    worksheet = (
        writer.sheets["Sheet1"]
        if "Sheet1" in writer.sheets
        else writer.book.create_sheet("Sheet1")
    )
    data = []
    
    # Title
    schema_version = yaml_data.get("version", '')
    model = yaml_data.get("model", {})
    revision = model.get('revision','')
    revision_date = model.get('revision_date', '')
    
    title_font = Font(size=20, bold=True, name="Arial")
    
    cell = worksheet["A1"]
    cell.value = f"Geology Data Model"
    cell.font = title_font
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    worksheet.row_dimensions[current_row].height = 30
    
   
    cell = worksheet["A2"]
    cell.value = "Revision"
    cell.font = Font(size=12, bold=True, name="Arial")
    cell = worksheet["B2"]
    cell.value = revision
    cell.font = Font(size=12, bold=True, name="Arial")
    
    cell = worksheet["A3"]
    cell.value = "Date"
    cell.font = Font(size=12, bold=True, name="Arial")
    cell = worksheet["B3"]
    cell.value = revision_date
    cell.font = Font(size=12, bold=True, name="Arial")
    
    cell = worksheet["A4"]
    cell.value = "Schema"
    cell.font = Font(size=12, bold=True, name="Arial")
    cell = worksheet["B4"]
    cell.value = schema_version
    cell.font = Font(size=12, bold=True, name="Arial")
    
    
    

    # Format the headers (bold, size 14, center alignment)
    current_row += 4
    
    worksheet[f"C{current_row}"] = "Deutsch"
    worksheet[f"D{current_row}"] = "Français"
    worksheet[f"E{current_row}"] = "Type"
    worksheet[f"F{current_row}"] = "Mandatory"
    worksheet[f"G{current_row}"] = "Cardinality"
    
    
    header_font = Font(size=14, bold=True, color="dddddd", name="Arial")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(
        start_color="333333", end_color="dddddd", fill_type="solid"
    )
    
    
    
   
    

    # Apply the formatting to the headers
    for col in [chr(i) for i in range(ord("A"), ord("G") + 1)]:
        cell = worksheet[f"{col}{current_row}"]
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        
    current_row += 1

    for theme in yaml_data.get("themes", []):
        theme_name = theme.get("name", "")
        #theme_desc_de = theme.get("description", {}).get("de", "")
        #theme_desc_fr = theme.get("description", {}).get("fr", "")
        for cls in theme.get("classes", []):
            class_name = cls.get("name", "")
            if "Unconsolidated_Deposits_PLG" in class_name:
                print(cls.get("description", {}))
            abrev = cls.get("abrev", "")
            table = cls.get("table", "")
            description_de = cls.get("description", {}).get("de", "")
            description_fr = cls.get("description", {}).get("fr", "")
            print(
                description_de.strip()
                if description_de != previous_description_de
                else ""
            )

            # Iterate over attributes
            for attr in cls.get("attributes", []):
                attr_name = attr.get("name", "")
                attr_type = attr.get("att_type", "")
                attr_value = attr.get("value", "")
                attr_desc_de = attr.get("description", {}).get("de", "")
                attr_desc_fr = attr.get("description", {}).get("fr", "")
                cardinality = attr.get("cardinality", "")
                mandatory = attr.get("mandatory", False)

                (
                    theme_name if theme_name != previous_theme else "",
                )  # Omit repeating theme name
                (
                    class_name if class_name != previous_class else "",
                )  # Omit repeating class name
                (description_de if description_de != previous_description_de else "",)
                (description_fr if description_fr != previous_description_fr else "",)

                # Prepare the row data
                row_data = [
                    # "Abbreviation": abrev if abrev != previous_abrev else "",  # Omit repeating class name
                    # "Table Name": table if table != previous_table else "",  # Omit repeating class name,
                    attr_name,
                    attr_desc_de,
                    attr_desc_fr,
                    attr_type,
                    attr_value,
                    "yes" if mandatory else "no",
                    f"[{cardinality}]",
                ]
                data.append(row_data)

                

                # Theme
                if theme_name != previous_theme:
                  cell = worksheet.cell(row=current_row, column=1, value="Theme")
                  cell.font = Font(bold=True, name="Arial")
                  cell = worksheet.cell(row=current_row, column=2, value=theme_name)
                  cell.font = font = Font(size=20, bold=True, name="Arial")
                  
               
                  
                         
                # Classe
                cell = worksheet.cell(row=current_row + 1, column=1, value="Class")
                cell.font = Font(bold=True, name="Arial")

                
                cell = worksheet.cell(row=current_row + 1, column=2, value=class_name)
                cell.font = Font(size=14, bold=True, name="Arial")

                my_align = Alignment(wrapText=True, vertical="top")
                for col_range in range(1, 12):
                    cell_title = worksheet.cell(current_row + 1, col_range)
                    cell_title.fill = PatternFill(
                        start_color="dddddd", end_color="dddddd", fill_type="solid"
                    )
                worksheet.cell(row=current_row + 2, column=3, value=description_de)
                worksheet.cell(row=current_row + 2, column=4, value=description_fr)
                
                worksheet.cell(row=current_row + 2, column=3).alignment = my_align
                worksheet.cell(row=current_row + 2, column=4).alignment = my_align
                cell = worksheet.cell(row=current_row + 2, column=1, value="Description")
                cell.font = Font(bold=True, name="Arial")
                cell.alignment = my_align
                
                worksheet.cell(row=current_row + 3, column=2, value=abrev)
                cell = worksheet.cell(row=current_row + 3, column=1, value="Abrev")
                cell.font = Font(bold=True, name="Arial")
                
                worksheet.cell(row=current_row + 4, column=2, value=table)
                cell = worksheet.cell(row=current_row + 4, column=1, value="Table")
                cell.font = Font(bold=True, name="Arial")
                
                cell = worksheet.cell(row=current_row + 5, column=1, value="Attributes")
                cell.font = Font(bold=True, name="Arial")

                

                worksheet.row_dimensions[current_row].height = 30
                worksheet.row_dimensions[current_row + 1].height = 20

                worksheet.row_dimensions[current_row + 2].height = 100

                column_index = 0

                current_row += 5

                for attr in data:
                    (
                        attr_name,
                        attr_desc_de,
                        attr_desc_fr,
                        attr_type,
                        attr_value,
                        mandatory,
                        cardinality,
                    ) = attr

                    worksheet.cell(
                        row=current_row, column=2, value=attr_name.upper()
                    )  # Indented description
                    worksheet.cell(
                        row=current_row, column=3, value=attr_desc_de
                    ).alignment = my_align  # Indented cardinality
                    worksheet.cell(
                        row=current_row, column=4, value=attr_desc_fr
                    ).alignment = my_align  # Indented mandatory

                    worksheet.cell(row=current_row, column=5, value=attr_type)
                    worksheet.cell(row=current_row, column=6, value=mandatory)
                    worksheet.cell(row=current_row, column=7, value=cardinality)
                    current_row += 1  # Move down for each attribute

                current_row += 2  # Next class

                dim_holder = DimensionHolder(worksheet=worksheet)

                for col in range(worksheet.min_column, worksheet.max_column + 1):
                    column_name = get_column_letter(col)
                    width = 20
                    if column_name.upper() in ["C", "D"]:
                        width = 75
                    if column_name.upper() in ["E", "F", "G"]:
                        width = 20

                    dim_holder[column_name] = ColumnDimension(
                       worksheet, min=col, max=col, width=width
                    )
                    worksheet.column_dimensions = dim_holder
                
                # Update previous_theme and previous_class
                previous_theme = theme_name
                previous_class = class_name
                previous_abrev = abrev
                previous_table = table
                previous_description_de = description_de
                previous_description_fr = description_fr    
            
