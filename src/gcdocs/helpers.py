import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Optional

from loguru import logger


def write_output(
    data: dict,
    output_path: Path,
    format: str,
    with_metadata: bool = False,
    metadata: dict = None,
):
    """
    Write extracted data to file in the specified format.

    Args:
        data: Dict of {code: label} pairs
        output_path: Output file path
        format: Output format (json, csv, xlsx)
        with_metadata: Include metadata in output (JSON only)
        metadata: Metadata dict to include
    """
    if format == "json":
        if with_metadata and metadata:
            output_data = {"_metadata": metadata, "codes": data}
        else:
            output_data = data

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False)

    elif format == "csv":
        with open(output_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(["code", "label"])
            for code, label in data.items():
                writer.writerow([code, label])

    elif format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise click.ClickException(
                "openpyxl is required for XLSX export. Install with: pip install openpyxl"
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Codes"

        # Header style
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
        )

        # Write headers
        ws["A1"] = "code"
        ws["B1"] = "label"
        ws["A1"].font = header_font
        ws["B1"].font = header_font
        ws["A1"].fill = header_fill
        ws["B1"].fill = header_fill

        # Write data
        for row, (code, label) in enumerate(data.items(), start=2):
            ws[f"A{row}"] = code
            ws[f"B{row}"] = label

        # Auto-adjust column widths
        for col in ["A", "B"]:
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell = ws[f"{col}{row}"]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col].width = min(max_length + 2, 80)

        # Add metadata sheet if requested
        if with_metadata and metadata:
            ws_meta = wb.create_sheet(title="Metadata")
            for row, (key, value) in enumerate(metadata.items(), start=1):
                ws_meta[f"A{row}"] = key
                ws_meta[f"B{row}"] = str(value) if value is not None else ""

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(output_path)


def extract_coded_domains(
    data: dict, prefix: str = "GC_", sorted_output: bool = False
) -> dict:
    """
    Extract coded values from coded_domain entries.

    Handles two ESRI export formats:

    Format 1 (simplified/transformed):
        {"coded_domain": {"DOMAIN_NAME": {"codedValues": {"code": "label"}}}}

    Format 2 (native ESRI export):
        {"domains": [{"type": "codedValue", "name": "DOMAIN_NAME",
                      "codedValues": [{"name": "label", "code": 123}]}]}

    Args:
        data: The full SDE schema dictionary
        prefix: Only include domains starting with this prefix (empty string = all)
        sorted_output: Sort by numeric key if True

    Returns:
        Dict of {code: label} pairs
    """
    result = {}

    # Format 1: coded_domain dict (simplified/transformed format)
    coded_domain = data.get("coded_domain", {})
    if isinstance(coded_domain, dict):
        for domain_name, domain_data in coded_domain.items():
            if prefix and not domain_name.startswith(prefix):
                continue

            coded_values = domain_data.get("codedValues", {})
            # Handle dict format: {"code": "label"}
            if isinstance(coded_values, dict):
                for code, label in coded_values.items():
                    result[str(code)] = label
            # Handle list format: [{"code": x, "name": y}]
            elif isinstance(coded_values, list):
                for cv in coded_values:
                    if "code" in cv and "name" in cv:
                        result[str(cv["code"])] = cv["name"]

    # Format 2: domains array (native ESRI export format)
    domains = data.get("domains", [])
    if isinstance(domains, list):
        for domain in domains:
            # Skip range domains and other non-codedValue types
            if domain.get("type") != "codedValue":
                continue

            domain_name = domain.get("name", "")
            if prefix and not domain_name.startswith(prefix):
                continue

            coded_values = domain.get("codedValues", [])
            if isinstance(coded_values, list):
                for cv in coded_values:
                    if "code" in cv and "name" in cv:
                        result[str(cv["code"])] = cv["name"]
            elif isinstance(coded_values, dict):
                for code, label in coded_values.items():
                    result[str(code)] = label

    if sorted_output:
        try:
            result = dict(sorted(result.items(), key=lambda x: int(x[0])))
        except ValueError:
            result = dict(sorted(result.items()))

    return result


def extract_subtypes_from_schema(data: dict, sorted_output: bool = False) -> dict:
    """
    Extract subtypes from SDE schema.

    Handles multiple formats:
    - Simple dict format: {"subtypes": {"code": "name", ...}}
    - Nested list format: {"subtypes": [{"subtypeCode": x, "subtypeName": y}, ...]}
    - ESRI native format in featureClasses/tables

    Args:
        data: The full SDE schema dictionary
        sorted_output: Sort by numeric key if True

    Returns:
        Dict of {subtype_code: subtype_name} pairs
    """
    result = {}

    # Check top-level subtypes
    subtypes_data = data.get("subtypes", {})

    # Simple dict format
    if isinstance(subtypes_data, dict):
        result = {str(k): v for k, v in subtypes_data.items()}
    # List format with subtypeCode/subtypeName
    elif isinstance(subtypes_data, list):
        for subtype in subtypes_data:
            if "subtypeCode" in subtype and "subtypeName" in subtype:
                result[str(subtype["subtypeCode"])] = subtype["subtypeName"]

    # Also search recursively for nested subtypes in featclasses/tables/featureClasses
    def find_nested_subtypes(obj):
        if isinstance(obj, dict):
            # Check for subtypes in this dict
            if "subtypes" in obj:
                subtypes = obj["subtypes"]
                if isinstance(subtypes, list):
                    for subtype in subtypes:
                        if "subtypeCode" in subtype and "subtypeName" in subtype:
                            result[str(subtype["subtypeCode"])] = subtype["subtypeName"]
                elif isinstance(subtypes, dict):
                    for code, name in subtypes.items():
                        if isinstance(name, str):
                            result[str(code)] = name

            # Recurse into all values
            for value in obj.values():
                find_nested_subtypes(value)
        elif isinstance(obj, list):
            for item in obj:
                find_nested_subtypes(item)

    find_nested_subtypes(data)

    if sorted_output:
        try:
            result = dict(sorted(result.items(), key=lambda x: int(x[0])))
        except ValueError:
            result = dict(sorted(result.items()))

    return result
