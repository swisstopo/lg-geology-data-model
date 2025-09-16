"""
Excel exporter for gcdocs
Based on the model.py export functionality from original code
"""

import pandas as pd
from pathlib import Path
from loguru import logger
import yaml
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from gcdocs.translation.model_translator  import ModelTranslationManager


class EnhancedXLSXExporter:
    """
    Enhanced XLSX exporter for translation workflow
    Extends your existing XLSXExporter
    """

    def __init__(self, model_translation_manager: ModelTranslationManager):
        self.tm = model_translation_manager

    def export_for_translation(self, yaml_file: str, output_file: str):
        """Export YAML to translation-ready Excel format"""

        # Load YAML data
        with open(yaml_file, "r", encoding="utf-8") as f:
            yaml_data = yaml.safe_load(f)

        # Create workbook with styling
        wb = Workbook()
        ws = wb.active
        ws.title = "Model Translations"

        # Define styles (similar to your existing xlsx.py styling)
        header_font = Font(name="Arial", bold=True, size=12)
        theme_font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
        class_font = Font(name="Arial", bold=True, size=11)
        theme_fill = PatternFill(
            start_color="2F4F4F", end_color="2F4F4F", fill_type="solid"
        )
        class_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )

        # Headers
        headers = [
            "Context",
            "Item",
            "Key",
            "Deutsch",
            "Français",
            "Italiano",
            "English",
            "Type",
            "Mandatory",
            "Cardinality",
            "Notes",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font

        current_row = 2

        # Model metadata
        model_info = yaml_data.get("model", {})
        ws.cell(row=current_row, column=1, value="Metadata").font = header_font
        ws.cell(row=current_row, column=2, value="Revision")
        ws.cell(row=current_row, column=4, value=model_info.get("revision", ""))
        current_row += 1

        ws.cell(row=current_row, column=2, value="Date")
        ws.cell(
            row=current_row, column=4, value=str(model_info.get("revision_date", ""))
        )
        current_row += 2

        # Process themes and classes
        for theme in yaml_data.get("themes", []):
            theme_name = theme.get("name", "")
            theme_key = self._make_key(theme_name)

            # Theme row
            self._add_theme_row(ws, current_row, theme_name, theme_fill, theme_font)
            current_row += 1

            # Process classes in theme
            for cls in theme.get("classes", []):
                class_name = cls.get("name", "")
                class_key = self._make_key(class_name)

                # Class row
                self._add_class_row(ws, current_row, class_name, class_fill, class_font)
                current_row += 1

                # Class description
                desc_key = f"theme.{theme_key}.class.{class_key}.description"
                current_row = self._add_description_row(
                    ws, current_row, "Description", desc_key, cls.get("description", {})
                )

                # Abbreviation and Table (non-translatable metadata)
                ws.cell(row=current_row, column=1, value="Metadata")
                ws.cell(row=current_row, column=2, value="Abrev")
                ws.cell(row=current_row, column=4, value=cls.get("abrev", ""))
                current_row += 1

                ws.cell(row=current_row, column=1, value="Metadata")
                ws.cell(row=current_row, column=2, value="Table")
                ws.cell(row=current_row, column=4, value=cls.get("table", ""))
                current_row += 1

                # Attributes header
                ws.cell(row=current_row, column=1, value="Attributes").font = class_font
                current_row += 1

                # Process attributes
                for attr in cls.get("attributes", []):
                    attr_name = attr.get("name", "").upper()
                    attr_key = f"theme.{theme_key}.class.{class_key}.attr.{attr_name.lower()}.description"

                    # Attribute row
                    ws.cell(row=current_row, column=1, value="Attribute")
                    ws.cell(row=current_row, column=2, value=attr_name)
                    ws.cell(row=current_row, column=3, value=attr_key)

                    # Add existing translations
                    attr_desc = attr.get("description", {})
                    for i, lang in enumerate(["de", "fr", "it", "en"]):
                        col = 4 + i
                        existing = self._get_existing_translation(
                            attr_key, lang, attr_desc
                        )
                        ws.cell(row=current_row, column=col, value=existing)

                    # Add metadata columns
                    ws.cell(row=current_row, column=8, value=attr.get("att_type", ""))
                    ws.cell(
                        row=current_row,
                        column=9,
                        value="yes" if attr.get("mandatory", False) else "no",
                    )
                    ws.cell(
                        row=current_row,
                        column=10,
                        value=f"[{attr.get('cardinality', '')}]",
                    )

                    current_row += 1

                current_row += 1  # Space between classes

        # Format columns
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            if col == 3:  # Key column
                ws.column_dimensions[column_letter].width = 50
            elif 4 <= col <= 7:  # Language columns
                ws.column_dimensions[column_letter].width = 60
            else:
                ws.column_dimensions[column_letter].width = 15

        # Enable text wrapping for description columns
        for row in ws.iter_rows(min_row=2):
            for cell in row[3:7]:  # Language columns
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        wb.save(output_file)
        logger.info(f"Exported translation template to {output_file}")

    def _make_key(self, name: str) -> str:
        """Convert name to key format"""
        return name.lower().replace(" ", "_").replace("-", "_")

    def _add_theme_row(self, ws, row, theme_name, fill, font):
        """Add theme row with styling"""
        cell = ws.cell(row=row, column=1, value="Theme")
        cell.font = font
        cell.fill = fill

        cell = ws.cell(row=row, column=2, value=theme_name)
        cell.font = font
        cell.fill = fill

        for col in range(2, 10):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = fill

    def _add_class_row(self, ws, row, class_name, fill, font):
        """Add class row with styling"""
        cell = ws.cell(row=row, column=1, value="Class")
        cell.font = font
        cell.fill = fill

        cell = ws.cell(row=row, column=2, value=class_name)
        cell.font = font
        cell.fill = fill

        for col in range(2, 10):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = fill

    def _add_description_row(self, ws, row, context, key, description_dict):
        """Add description row with translations"""
        ws.cell(row=row, column=1, value=context)
        ws.cell(row=row, column=2, value="Description")
        ws.cell(row=row, column=3, value=key)

        # Add existing translations
        for i, lang in enumerate(["de", "fr", "it", "en"]):
            col = 4 + i
            existing = self._get_existing_translation(key, lang, description_dict)
            ws.cell(row=row, column=col, value=existing)

        return row + 1

    def _get_existing_translation(
        self, key: str, lang: str, description_dict: dict
    ) -> str:
        """Get existing translation from various sources"""
        # First check translation manager
        if key in self.tm.translations.get(lang, {}):
            return self.tm.translations[lang][key]

        # Then check embedded description dict (for migration)
        if isinstance(description_dict, dict) and lang in description_dict:
            return description_dict[lang]

        return ""

    def import_from_excel(self, excel_file: str) -> dict:
        """Import completed translations from Excel"""
        df = pd.read_excel(excel_file)

        # Clear existing model translations
        for lang in self.tm.supported_languages:
            self.tm.translations[lang] = {}

        # Process rows
        for _, row in df.iterrows():
            context = str(row.get("Context", "")).strip()
            if context not in ["Theme", "Class", "Attribute", "Description"]:
                continue

            key = str(row.get("Key", "")).strip()
            if not key:
                continue

            # Store translations
            for lang in ["de", "fr", "it", "en"]:
                lang_col = lang.title()  # Deutsch, Français, etc.
                col_name = {
                    "de": "Deutsch",
                    "fr": "Français",
                    "it": "Italiano",
                    "en": "English",
                }[lang]

                if (
                    col_name in row
                    and pd.notna(row[col_name])
                    and str(row[col_name]).strip()
                ):
                    self.tm.translations[lang][key] = str(row[col_name]).strip()

        # Save translation files
        self.tm.save_translations()
        logger.info("Imported model translations from Excel")

        return self._build_yaml_structure_with_keys(df)

    def _build_yaml_structure_with_keys(self, df: pd.DataFrame) -> dict:
        """Build YAML structure using translation keys instead of embedded text"""
        # This would rebuild the YAML with description_key references
        # Implementation depends on your exact YAML structure requirements
        pass


class XLSXExporter:
    """Export datamodel to Excel format"""

    def export(self, yaml_file: str, output_file: str):
        """Export YAML datamodel to Excel"""
        # Load YAML data
        with open(yaml_file, 'r', encoding='utf-8') as f:
            yaml_data = yaml.safe_load(f)

        # Convert to Excel-friendly format
        excel_data = self._prepare_excel_data(yaml_data)

        # Write to Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            self._write_excel_sheets(excel_data, writer)

        logger.info(f"Exported datamodel to {output_file}")

    def _prepare_excel_data(self, yaml_data: Dict[str, Any]) -> Dict[str, List[Dict]]:
        """Convert YAML data to Excel-friendly format"""
        excel_data = {
            'Metadata': [],
            'Themes': [],
            'Classes': [],
            'Attributes': []
        }

        # Metadata sheet
        metadata = yaml_data.get('model', {})
        excel_data['Metadata'] = [
            {'Key': 'Version', 'Value': yaml_data.get('version', '')},
            {'Key': 'Date', 'Value': yaml_data.get('date', '')},
            {'Key': 'Revision', 'Value': metadata.get('revision', '')},
            {'Key': 'Revision Date', 'Value': metadata.get('revision_date', '')},
        ]

        # Process themes and classes
        for theme in yaml_data.get('themes', []):
            theme_name = theme.get('name', '')

            # Add theme
            excel_data['Themes'].append({
                'Name': theme_name,
                'Description_DE': theme.get('description', {}).get('de', ''),
                'Description_FR': theme.get('description', {}).get('fr', '')
            })

            # Process classes
            for cls in theme.get('classes', []):
                class_name = cls.get('name', '')

                # Add class
                excel_data['Classes'].append({
                    'Theme': theme_name,
                    'Name': class_name,
                    'Abbreviation': cls.get('abrev', ''),
                    'Table': cls.get('table', ''),
                    'Description_DE': cls.get('description', {}).get('de', ''),
                    'Description_FR': cls.get('description', {}).get('fr', '')
                })

                # Process attributes
                for attr in cls.get('attributes', []):
                    excel_data['Attributes'].append({
                        'Theme': theme_name,
                        'Class': class_name,
                        'Name': attr.get('name', ''),
                        'Alias': attr.get('alias', ''),
                        'Type': attr.get('att_type', ''),
                        'Value': attr.get('value', ''),
                        'Description_DE': attr.get('description', {}).get('de', ''),
                        'Description_FR': attr.get('description', {}).get('fr', ''),
                        'Cardinality': attr.get('cardinality', ''),
                        'Mandatory': 'Yes' if attr.get('mandatory') else 'No',
                        'Status': attr.get('status', ''),
                        'Change': attr.get('change', '')
                    })

        return excel_data

    def _write_excel_sheets(self, excel_data: Dict[str, List[Dict]], writer: pd.ExcelWriter):
        """Write data to Excel sheets with formatting"""
        for sheet_name, data in excel_data.items():
            if not data:
                continue

            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Get worksheet for formatting
            worksheet = writer.sheets[sheet_name]

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Max width of 50
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Style headers
            if len(df) > 0:
                for cell in worksheet[1]:  # First row (headers)
                    cell.font = writer.book.active.parent._fonts[0].copy(bold=True)
                    cell.fill = writer.book.active.parent._fills[0].copy()


class DataModelImporter:
    """Import Excel back to YAML (for round-trip functionality)"""

    def from_excel(self, excel_file: str, output_yaml: str):
        """Convert Excel file back to YAML datamodel"""
        # Read Excel sheets
        excel_data = pd.read_excel(excel_file, sheet_name=None)

        # Convert back to YAML structure
        yaml_data = self._excel_to_yaml(excel_data)

        # Write YAML file
        with open(output_yaml, 'w', encoding='utf-8') as f:
            yaml.dump(yaml_data, f, default_flow_style=False, allow_unicode=True)

        logger.info(f"Imported Excel to {output_yaml}")

    def _excel_to_yaml(self, excel_data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """Convert Excel data back to YAML structure"""
        yaml_data = {}

        # Process metadata
        if 'Metadata' in excel_data:
            metadata_df = excel_data['Metadata']
            for _, row in metadata_df.iterrows():
                key = row['Key']
                value = row['Value']

                if key == 'Version':
                    yaml_data['version'] = value
                elif key == 'Date':
                    yaml_data['date'] = value
                elif key in ['Revision', 'Revision Date']:
                    if 'model' not in yaml_data:
                        yaml_data['model'] = {}
                    yaml_data['model'][key.lower().replace(' ', '_')] = value

        # Process themes and classes
        themes = {}

        if 'Classes' in excel_data:
            classes_df = excel_data['Classes']

            for _, row in classes_df.iterrows():
                theme_name = row['Theme']

                if theme_name not in themes:
                    themes[theme_name] = {
                        'name': theme_name,
                        'classes': []
                    }

                class_data = {
                    'name': row['Name'],
                    'abrev': row['Abbreviation'],
                    'table': row['Table'],
                    'description': {
                        'de': row['Description_DE'],
                        'fr': row['Description_FR']
                    },
                    'attributes': []
                }

                themes[theme_name]['classes'].append(class_data)

        # Process attributes
        if 'Attributes' in excel_data:
            attributes_df = excel_data['Attributes']

            # Group attributes by theme and class
            for _, row in attributes_df.iterrows():
                theme_name = row['Theme']
                class_name = row['Class']

                # Find the corresponding class
                theme = themes.get(theme_name)
                if not theme:
                    continue

                target_class = None
                for cls in theme['classes']:
                    if cls['name'] == class_name:
                        target_class = cls
                        break

                if not target_class:
                    continue

                # Add attribute
                attr_data = {
                    'name': row['Name'],
                    'att_type': row['Type'],
                    'description': {
                        'de': row['Description_DE'],
                        'fr': row['Description_FR']
                    },
                    'cardinality': row['Cardinality'],
                    'mandatory': row['Mandatory'].lower() == 'yes'
                }

                # Add optional fields
                if pd.notna(row.get('Alias')):
                    attr_data['alias'] = row['Alias']
                if pd.notna(row.get('Value')):
                    attr_data['value'] = row['Value']
                if pd.notna(row.get('Status')):
                    attr_data['status'] = row['Status']
                if pd.notna(row.get('Change')):
                    attr_data['change'] = row['Change']

                target_class['attributes'].append(attr_data)

        # Convert themes dict to list
        yaml_data['themes'] = list(themes.values())

        return yaml_data