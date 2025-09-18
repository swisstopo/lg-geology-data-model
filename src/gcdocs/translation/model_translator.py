# Enhanced Translation System for gcdocs
# Extends existing geological code translations with hierarchical YAML translations

import json
import logging
import os
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import yaml
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from loguru import logger

# Update your config.py to include the new languages
AVAILABLE_LANGUAGES = ["de", "fr", "it", "en"]  # Extended from your current DE/FR

NB_COLUMNS = 11


class ModelTranslationManager:
    """
    Manages translations for YAML model descriptions (themes, classes, attributes)
    Works alongside existing SimpleTranslator for geological codes
    """

    def __init__(self, translations_dir: str = "translations"):
        self.translations_dir = Path(translations_dir)
        self.supported_languages = AVAILABLE_LANGUAGES
        self.translations = {}
        self.load_translations()

    def load_translations(self):
        """Load hierarchical translation files"""
        for lang in self.supported_languages:
            file_path = self.translations_dir / f"model_{lang}.json"
            if file_path.exists():
                with open(file_path, "r", encoding="utf-8") as f:
                    self.translations[lang] = json.load(f)
            else:
                self.translations[lang] = {}
                if file_path.parent.exists():  # Only warn if directory exists
                    logger.debug(
                        f"Model translation file for {lang} not found: {file_path}"
                    )

    def get_translation(self, key: str, lang: str, fallback: str = "") -> str:
        """Get hierarchical translation for a key"""
        if lang in self.translations and key in self.translations[lang]:
            return self.translations[lang][key]
        elif fallback:
            return fallback
        else:
            logger.warning(f"Missing model translation for key '{key}' in {lang}")
            return f"[MISSING: {key}]"

    def save_translations(self):
        """Save all translation files"""
        self.translations_dir.mkdir(exist_ok=True)

        for lang in self.supported_languages:
            file_path = self.translations_dir / f"model_{lang}.json"
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(
                    self.translations[lang],
                    f,
                    ensure_ascii=False,
                    indent=2,
                    sort_keys=True,
                )


class ModelProcessor:
    """
    Processes YAML models with translation support
    Works with your existing MarkdownGenerator
    """

    def __init__(self, model_tm: ModelTranslationManager, geol_translator):
        self.model_tm = model_tm
        self.geol_translator = (
            geol_translator  # Your existing geological code translator
        )

    def migrate_yaml_to_keys(self, yaml_data: dict) -> dict:
        """Migrate existing YAML to use translation keys"""

        def process_themes(themes):
            new_themes = []
            for theme in themes:
                theme_name = theme.get("name", "")
                theme_key = self._make_key(theme_name)
                new_theme = {"name": theme_name, "classes": []}

                for cls in theme.get("classes", []):
                    class_name = cls.get("name", "")
                    class_key = self._make_key(class_name)

                    new_class = {
                        "name": cls["name"],
                        "abrev": cls.get("abrev", ""),
                        "table": cls.get("table", ""),
                        "attributes": [],
                    }

                    # Handle class description
                    if "description" in cls:
                        desc_key = f"theme.{theme_key}.class.{class_key}.description"
                        new_class["description_key"] = desc_key

                        # Store translations
                        if isinstance(cls["description"], dict):
                            for lang in ["de", "fr"]:
                                if (
                                    lang in cls["description"]
                                    and cls["description"][lang]
                                ):
                                    if lang not in self.model_tm.translations:
                                        self.model_tm.translations[lang] = {}
                                    self.model_tm.translations[lang][desc_key] = cls[
                                        "description"
                                    ][lang]

                    # Handle attributes
                    for attr in cls.get("attributes", []):
                        attr_name = attr.get("name", "").lower()
                        attr_key = f"theme.{theme_key}.class.{class_key}.attr.{attr_name}.description"

                        new_attr = {
                            "name": attr["name"],
                            "att_type": attr.get("att_type", ""),
                            "mandatory": attr.get("mandatory", False),
                            "cardinality": attr.get("cardinality", ""),
                        }

                        # Handle attribute description
                        if "description" in attr:
                            new_attr["description_key"] = attr_key

                            # Store translations
                            if isinstance(attr["description"], dict):
                                for lang in ["de", "fr"]:
                                    if (
                                        lang in attr["description"]
                                        and attr["description"][lang]
                                    ):
                                        if lang not in self.model_tm.translations:
                                            self.model_tm.translations[lang] = {}
                                        self.model_tm.translations[lang][attr_key] = (
                                            attr["description"][lang]
                                        )

                        # Copy other attribute properties
                        for key in ["value", "alias", "status", "change"]:
                            if key in attr:
                                new_attr[key] = attr[key]

                        new_class["attributes"].append(new_attr)

                    new_theme["classes"].append(new_class)

                new_themes.append(new_theme)

            return new_themes

        # Create migrated structure
        new_yaml = {
            "version": yaml_data.get("version", ""),
            "date": yaml_data.get("date", ""),
            "model": yaml_data.get("model", {}),
            "themes": process_themes(yaml_data.get("themes", [])),
            "annexes": yaml_data.get("annexes", []),
        }

        return new_yaml

    def render_with_translations(self, yaml_data: dict, language: str) -> dict:
        """Render YAML with translations for specified language"""

        def process_item(item):
            if isinstance(item, dict):
                new_item = {}
                for k, v in item.items():
                    if k.endswith("_key") and isinstance(v, str):
                        # Replace key with actual translation
                        field_name = k.replace("_key", "")
                        translation = self.model_tm.get_translation(v, language)
                        new_item[field_name] = translation
                    else:
                        new_item[k] = process_item(v)
                return new_item
            elif isinstance(item, list):
                return [process_item(sub_item) for sub_item in item]
            else:
                return item

        return process_item(yaml_data)

    def _make_key(self, name: str) -> str:
        """Convert name to key format"""
        return name.lower().replace(" ", "_").replace("-", "_")


class HierarchicalTranslationManager:
    """Manages hierarchical translations for the geology data model"""

    def __init__(self, translations_dir: str = "translations"):
        self.translations_dir = Path(translations_dir)
        self.supported_languages = ["de", "fr", "it", "en"]
        self.translations = {}
        self.load_translations()

    def load_translations(self):
        """Load all translation files"""
        for lang in self.supported_languages:
            file_path = self.translations_dir / f"{lang}.json"
            if file_path.exists():
                with open(file_path, "r", encoding="utf-8") as f:
                    self.translations[lang] = json.load(f)
            else:
                self.translations[lang] = {}
                logger.warning(f"Translation file for {lang} not found")

    def get_translation(
        self, key: str, target_lang: str, fallback_text: str = ""
    ) -> str:
        """
        Get translation with intelligent fallback logic

        Priority order:
        1. Target language (e.g., 'it')
        2. German ('de') - primary fallback
        3. French ('fr') - secondary fallback
        4. English ('en') - tertiary fallback
        5. Fallback text or [MISSING] indicator
        """

        # 1. Try target language first
        if target_lang in self.translations and key in self.translations[target_lang]:
            translation = self.translations[target_lang][key]
            if translation and translation.strip():
                return translation.strip()

        # 2. Try German fallback (most complete)
        if target_lang not in ["de", "it"] and key in self.translations.get("de", {}):
            german_text = self.translations["de"][key]
            if german_text and german_text.strip():
                logger.debug(f"Using German fallback for {key} ({target_lang})")
                return german_text.strip()

        # 3. Try French fallback
        if target_lang not in ["de", "fr"] and key in self.translations.get("fr", {}):
            french_text = self.translations["fr"][key]
            if french_text and french_text.strip():
                logger.debug(f"Using French fallback for {key} ({target_lang})")
                return french_text.strip()

        # 4. Try English fallback
        if target_lang != "en" and key in self.translations.get("en", {}):
            english_text = self.translations["en"][key]
            if english_text and english_text.strip():
                logger.debug(f"Using English fallback for {key} ({target_lang})")
                return english_text.strip()

        # 5. Use fallback text or show missing indicator
        if fallback_text and fallback_text.strip():
            logger.warning(
                f"No translation found for key '{key}' in {target_lang}, using fallback"
            )
            return fallback_text.strip()
        else:
            logger.error(f"Missing translation for key '{key}' in {target_lang}")
            return f"[MISSING {target_lang.upper()}: {key.split('.')[-1]}]"

    def get_translation_status(self) -> Dict[str, Dict[str, int]]:
        """Get translation completion statistics"""

        # Get all unique keys across all languages
        all_keys = set()
        for lang_dict in self.translations.values():
            all_keys.update(lang_dict.keys())

        stats = {}
        for lang in self.supported_languages:
            translated_keys = set(self.translations.get(lang, {}).keys())
            stats[lang] = {
                "translated": len(translated_keys),
                "total": len(all_keys),
                "percentage": round(len(translated_keys) / len(all_keys) * 100, 1)
                if all_keys
                else 0,
                "missing": len(all_keys - translated_keys),
            }

        return stats

    def get_missing_translations(self, target_lang: str) -> Dict[str, str]:
        """Get list of missing translations for a language with fallback text"""

        # Get all keys from other languages
        all_keys = set()
        for lang, translations in self.translations.items():
            if lang != target_lang:
                all_keys.update(translations.keys())

        # Find missing keys
        existing_keys = set(self.translations.get(target_lang, {}).keys())
        missing_keys = all_keys - existing_keys

        # Get fallback text for missing keys
        missing_with_fallback = {}
        for key in missing_keys:
            fallback = self.get_translation(key, "de")  # Use German as fallback source
            missing_with_fallback[key] = fallback

        return missing_with_fallback

    def export_to_structured_excel(self, yaml_data: dict, output_file: str):
        """Export YAML data to structured Excel format with translations"""

        # Create workbook with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Geology Data Model"

        # Define styles
        header_font = Font(bold=True, size=12, name="Arial")
        theme_font = Font(bold=True, size=14, color="FFFFFF", name="Arial")
        class_font = Font(bold=True, size=11, name="Arial")
        theme_fill = PatternFill(
            start_color="2F4F4F", end_color="2F4F4F", fill_type="solid"
        )
        class_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )

        # Header row
        headers = [
            "Type",
            "Name/Attribute",
            "Deutsch",
            "Français",
            "Italiano",
            "English",
            "Att_Type",
            "Mandatory",
            "Cardinality",
            "Notes",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font

        # Model metadata
        current_row = 2
        model_info = yaml_data.get("model", {})
        ws.cell(row=current_row, column=1, value="Revision").font = header_font
        ws.cell(row=current_row, column=2, value=model_info.get("revision", ""))
        current_row += 1

        ws.cell(row=current_row, column=1, value="Date").font = header_font
        ws.cell(row=current_row, column=2, value=model_info.get("revision_date", ""))
        current_row += 1

        ws.cell(row=current_row, column=1, value="Schema").font = header_font
        ws.cell(row=current_row, column=2, value=yaml_data.get("version", ""))
        current_row += 2

        # Process themes and classes
        for theme in yaml_data.get("themes", []):
            theme_name = theme.get("name", "")

            # Theme row
            cell = ws.cell(row=current_row, column=1, value="Theme")
            cell.font = theme_font
            cell.fill = theme_fill

            cell = ws.cell(row=current_row, column=2, value=theme_name)
            cell.font = theme_font
            cell.fill = theme_fill

            for col in range(2, NB_COLUMNS):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = theme_fill

            current_row += 1

            # Process classes in theme
            for cls in theme.get("classes", []):
                class_name = cls.get("name", "")
                theme_key = theme_name.lower().replace("_", "_").replace(" ", "_")
                class_key = class_name.lower().replace("_", "_")

                # Class row
                cell = ws.cell(row=current_row, column=1, value="Class")
                cell.font = class_font
                cell.fill = class_fill

                cell = ws.cell(row=current_row, column=2, value=class_name)
                cell.font = class_font
                cell.fill = class_fill

                for col in range(2, NB_COLUMNS):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = class_fill

                current_row += 1

                # Class description
                desc_key = f"theme.{theme_key}.class.{class_key}.description"
                cell = ws.cell(row=current_row, column=1, value="Description")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.row_dimensions[current_row].height = 60  # TODO make it dynamic

                ws.cell(
                    row=current_row,
                    column=3,
                    value=self._get_existing_translation(
                        desc_key, "de", cls.get("description", {})
                    ),
                )
                ws.cell(
                    row=current_row,
                    column=4,
                    value=self._get_existing_translation(
                        desc_key, "fr", cls.get("description", {})
                    ),
                )
                ws.cell(
                    row=current_row,
                    column=5,
                    value=self._get_existing_translation(
                        desc_key, "it", cls.get("description", {})
                    ),
                )
                ws.cell(
                    row=current_row,
                    column=6,
                    value=self._get_existing_translation(
                        desc_key, "en", cls.get("description", {})
                    ),
                )
                current_row += 1

                # Abbreviation
                ws.cell(row=current_row, column=1, value="Abrev")
                ws.cell(row=current_row, column=2, value=cls.get("abrev", ""))
                current_row += 1

                # Table
                ws.cell(row=current_row, column=1, value="Table")
                ws.cell(row=current_row, column=2, value=cls.get("table", ""))
                current_row += 1

                # Attributes header
                ws.cell(row=current_row, column=1, value="Attributes").font = class_font
                current_row += 1

                # Process attributes
                for attr in cls.get("attributes", []):
                    attr_name = attr.get("name", "").upper()
                    attr_key = (
                        f"theme.{theme_key}.class.{class_key}.attr.{attr_name.lower()}"
                    )

                    # Attribute row
                    ws.cell(row=current_row, column=2, value=attr_name)
                    ws.cell(
                        row=current_row,
                        column=3,
                        value=self._get_existing_translation(
                            f"{attr_key}.description", "de", attr.get("description", {})
                        ),
                    )
                    ws.cell(
                        row=current_row,
                        column=4,
                        value=self._get_existing_translation(
                            f"{attr_key}.description", "fr", attr.get("description", {})
                        ),
                    )
                    ws.cell(
                        row=current_row,
                        column=5,
                        value=self._get_existing_translation(
                            f"{attr_key}.description", "it", attr.get("description", {})
                        ),
                    )
                    ws.cell(
                        row=current_row,
                        column=6,
                        value=self._get_existing_translation(
                            f"{attr_key}.description", "en", attr.get("description", {})
                        ),
                    )
                    ws.cell(row=current_row, column=7, value=attr.get("att_type", ""))
                    ws.cell(
                        row=current_row,
                        column=8,
                        value="yes" if attr.get("mandatory", False) else "no",
                    )
                    ws.cell(
                        row=current_row,
                        column=9,
                        value=f"[{attr.get('cardinality', '')}]",
                    )

                    current_row += 1

                # Empty row between classes
                current_row += 1

        # Format columns
        for col in range(1, len(headers) + 1):
            column_letter = ws.cell(row=1, column=col).column_letter
            if col <= 2:  # Type and Name columns
                ws.column_dimensions[column_letter].width = 20
            elif col <= 6:  # Language columns
                ws.column_dimensions[column_letter].width = 60
            else:  # Other columns
                ws.column_dimensions[column_letter].width = 15

        # Enable text wrapping for description columns
        for row in ws.iter_rows(min_row=2):
            for cell in row[2:6]:  # Language columns
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        wb.save(output_file)
        print(f"Exported structured model to {output_file}")

    def _get_existing_translation(
        self, key: str, lang: str, description_dict: dict
    ) -> str:
        """Get existing translation from various sources"""
        # First check translation files
        if key in self.translations.get(lang, {}):
            return self.translations[lang][key]

        # Then check embedded description dict (for migration)
        if isinstance(description_dict, dict) and lang in description_dict:
            return description_dict[lang]

        return ""

    def import_from_structured_excel(self, input_file: str) -> dict:
        """Import translations from structured Excel file"""
        df = pd.read_excel(input_file, sheet_name=0)  # First sheet

        # Initialize translation dictionaries
        for lang in self.supported_languages:
            self.translations[lang] = {}

        current_theme = ""
        current_class = ""

        for _, row in df.iterrows():
            row_type = str(row.get("Type", "")).strip()
            name_attr = str(row.get("Name/Attribute", "")).strip()

            # Track current context
            if row_type == "Theme":
                current_theme = name_attr.lower().replace(" ", "_")
                continue
            elif row_type == "Class":
                current_class = name_attr.lower().replace(" ", "_")
                continue
            elif row_type == "Description" and current_theme and current_class:
                # Class description
                key = f"theme.{current_theme}.class.{current_class}.description"
                self._store_translations(key, row)
            elif (
                (row_type == "" or row_type == "nan")
                and name_attr
                and current_theme
                and current_class
            ):
                # Attribute
                attr_name = name_attr.lower()
                key = f"theme.{current_theme}.class.{current_class}.attr.{attr_name}.description"
                logger.debug(f"Attribute: {current_class} - {attr_name}")
                self._store_translations(key, row)
            else:
                logger.warning(f"Unkonw row type: {row_type}")

        # Save translation files
        self.save_translations()

        # Also return the processed data structure for YAML generation
        return self._build_yaml_structure(df)

    def _store_translations(self, key: str, row: pd.Series):
        """Store translations from Excel row"""
        lang_columns = {
            "de": "Deutsch",
            "fr": "Français",
            "it": "Italiano",
            "en": "English",
        }

        for lang, col_name in lang_columns.items():
            if (
                col_name in row
                and pd.notna(row[col_name])
                and str(row[col_name]).strip()
            ):
                self.translations[lang][key] = str(row[col_name]).strip()

    def _build_yaml_structure(self, df: pd.DataFrame) -> dict:
        """Rebuild YAML structure from Excel data"""
        yaml_structure = {
            "version": "",
            "model": {"revision": "", "revision_date": ""},
            "themes": [],
        }

        current_theme = None
        current_class = None

        for _, row in df.iterrows():
            row_type = str(row.get("Type", "")).strip()
            name_attr = str(row.get("Name/Attribute", "")).strip()

            if row_type == "Revision":
                yaml_structure["model"]["revision"] = name_attr
            elif row_type == "Date":
                yaml_structure["model"]["revision_date"] = name_attr
            elif row_type == "Schema":
                yaml_structure["version"] = name_attr
            elif row_type == "Theme":
                current_theme = {"name": name_attr, "classes": []}
                yaml_structure["themes"].append(current_theme)
            elif row_type == "Class":
                current_class = {
                    "name": name_attr,
                    "description_key": f"theme.{current_theme['name'].lower()}.class.{name_attr.lower()}.description",
                    "attributes": [],
                }
                current_theme["classes"].append(current_class)
            elif row_type == "Abrev":
                if current_class:
                    current_class["abrev"] = name_attr
            elif row_type == "Table":
                if current_class:
                    current_class["table"] = name_attr
            elif row_type == "" and name_attr and current_class:
                # Attribute
                attr = {
                    "name": name_attr.lower(),
                    "description_key": f"theme.{current_theme['name'].lower()}.class.{current_class['name'].lower()}.attr.{name_attr.lower()}.description",
                    "att_type": str(row.get("Att_Type", "")),
                    "mandatory": str(row.get("Mandatory", "")).lower() == "yes",
                    "cardinality": str(row.get("Cardinality", ""))
                    .replace("[", "")
                    .replace("]", ""),
                }
                current_class["attributes"].append(attr)

        return yaml_structure

    def save_translations(self):
        """Save all translation files"""
        self.translations_dir.mkdir(exist_ok=True)

        for lang in self.supported_languages:
            file_path = self.translations_dir / f"{lang}.json"
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(
                    self.translations[lang],
                    f,
                    ensure_ascii=False,
                    indent=2,
                    sort_keys=True,
                )


class HierarchicalDatamodelProcessor:
    """Process data model with hierarchical translation keys"""

    def __init__(self, translation_manager: HierarchicalTranslationManager):
        self.translation_manager = translation_manager

    def render_with_translations(self, yaml_data: dict, language: str) -> dict:
        """Render YAML data with translations for specified language"""

        def process_item(item, context_path=""):
            if isinstance(item, dict):
                new_item = {}
                for k, v in item.items():
                    if k.endswith("_key") and isinstance(v, str):
                        # Replace key with actual translation
                        field_name = k.replace("_key", "")
                        translation = self.translation_manager.get_translation(
                            v, language
                        )
                        new_item[field_name] = translation
                    else:
                        new_item[k] = process_item(v, context_path)
                return new_item
            elif isinstance(item, list):
                return [process_item(sub_item, context_path) for sub_item in item]
            else:
                return item

        return process_item(yaml_data)

    def migrate_from_embedded_yaml(self, yaml_data: dict) -> dict:
        """Migrate existing YAML with embedded translations to key-based system"""

        def process_themes(themes):
            new_themes = []
            for theme in themes:
                theme_name = theme.get("name", "").lower().replace(" ", "_")
                new_theme = {"name": theme["name"], "classes": []}

                for cls in theme.get("classes", []):
                    class_name = cls.get("name", "").lower().replace(" ", "_")
                    class_key = f"theme.{theme_name}.class.{class_name}"

                    new_class = {
                        "name": cls["name"],
                        "abrev": cls.get("abrev", ""),
                        "table": cls.get("table", ""),
                        "description_key": f"{class_key}.description",
                        "attributes": [],
                    }

                    # Store class description translations
                    if "description" in cls and isinstance(cls["description"], dict):
                        desc_key = f"{class_key}.description"
                        for lang in ["de", "fr"]:
                            if lang in cls["description"]:
                                if lang not in self.translation_manager.translations:
                                    self.translation_manager.translations[lang] = {}
                                self.translation_manager.translations[lang][
                                    desc_key
                                ] = cls["description"][lang]

                    # Process attributes
                    for attr in cls.get("attributes", []):
                        attr_name = attr.get("name", "").lower()
                        attr_key = f"{class_key}.attr.{attr_name}"

                        new_attr = {
                            "name": attr["name"],
                            "description_key": f"{attr_key}.description",
                            "att_type": attr.get("att_type", ""),
                            "mandatory": attr.get("mandatory", False),
                            "cardinality": attr.get("cardinality", ""),
                        }

                        # Store attribute description translations
                        if "description" in attr and isinstance(
                            attr["description"], dict
                        ):
                            for lang in ["de", "fr"]:
                                if lang in attr["description"]:
                                    if (
                                        lang
                                        not in self.translation_manager.translations
                                    ):
                                        self.translation_manager.translations[lang] = {}
                                    self.translation_manager.translations[lang][
                                        f"{attr_key}.description"
                                    ] = attr["description"][lang]

                        new_class["attributes"].append(new_attr)

                    new_theme["classes"].append(new_class)

                new_themes.append(new_theme)

            return new_themes

        # Create new structure
        new_yaml = {
            "version": yaml_data.get("version", ""),
            "date": yaml_data.get("date", ""),
            "model": yaml_data.get("model", {}),
            "themes": process_themes(yaml_data.get("themes", [])),
            "annexes": yaml_data.get("annexes", []),
        }

        return new_yaml
