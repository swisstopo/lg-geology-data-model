import sys
import textwrap
from collections import OrderedDict

import openpyxl
import pandas as pd
import yaml
from loguru import logger
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from ruamel.yaml import YAML


class Datamodel:
    def __init__(self):
        self.yaml_data = OrderedDict()
        self.yaml_data["metadata"] = {}
        self.yaml_data["themes"] = []

    def import_from_excel(self, file_path):
        """Import data from an Excel file into the yaml_data structure."""
        # Load the workbook and the first sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        yaml_data = OrderedDict()
        yaml_data["themes"] = []
        current_theme = None
        current_class = None

        n = 0
        revision = sheet["B2"].value
        revision_date = sheet["B3"].value
        schema_version = sheet["B3"].value

        yaml_data["model"] = {"revision": revision, "revision_date": revision_date}
        yaml_data["schema"] = str(schema_version)
        # Iterate over rows starting from the second row (skip headers)
        current_class_name = None
        current_class = None
        in_attributes = False
        for row in sheet.iter_rows(min_row=6, values_only=True):
            # Use the `or` operator to assign a default value (None or empty string) if the cell is empty

            if not any(row):
                continue

            n += 1
            theme_name = None
            class_name = None
            table_name = None
            abrev = None
            description_de = ""
            description_fr = ""
            attr_name = None
            attr_desc_de = attr_desc_fr = ""
            attr_type = mandatory = cardinality = None
            row_type = row[0] or ""
            if "Theme" in row_type:
                theme_name = row[1] or None
                in_attributes = False
            if "Class" in row_type:
                class_name = row[1] or None
                current_class_name = class_name
                in_attributes = False
            if "Description" in row_type:
                description_de = row[2] or ""
                description_fr = row[3] or ""
                in_attributes = False
            if "Abrev" in row_type:
                abrev = row[1] or None
                in_attributes = False
            if "Table" in row_type:
                table_name = row[1] or None
                in_attributes = False
            if "Attributes" in row_type:
                in_attributes = True
            if in_attributes:
                attr_name = row[1]
                attr_desc_de = row[2]
                attr_desc_fr = row[3]
                attr_type = row[4] or None
                mandatory = row[5] or "no"
                cardinality = row[6] or None

            # print(f"{row_type}, {attr_name}, in={in_attributes}")
            # print(row)
            # if n>20: return yaml_data

            # If a new theme is found
            if theme_name:
                # Create a new theme if it doesn't exist yet
                current_theme = {"name": theme_name, "classes": []}
                yaml_data["themes"].append(current_theme)

            # If a new class is found
            if class_name:
                # Create a new class if it doesn't exist yet
                current_class = {
                    "name": class_name,
                    "description": {
                        "de": description_de.strip(),
                        "fr": description_fr.strip(),
                    },
                    "attributes": [],
                }
                current_theme["classes"].append(current_class)
            if description_fr or description_de:
                current_class["description"]["de"] = "|" + description_de.strip()
                current_class["description"]["fr"] = "|" + description_fr.strip()
            if abrev:
                current_class["abrev"] = abrev
            if table_name:
                current_class["table"] = table_name

            # Add attribute to the current class
            if attr_name:
                attribute = {
                    "name": attr_name,
                    "att_type": attr_type,
                    "description": {"de": attr_desc_de, "fr": attr_desc_fr},
                    "mandatory": mandatory == "yes",
                    "cardinality": cardinality,
                }
                current_class["attributes"].append(attribute)
        self.yaml_data = yaml_data

        return yaml_data

    def export_to_excel(self, file_path):
        """Export the yaml_data structure to an Excel file."""

        yaml_data = self.yaml_data

        # Prepare the data for DataFrame
        data = []
        previous_theme = "dummy"
        previous_class = "dummy"
        previous_abrev = "dummy"
        previous_table = "dummy"
        previous_description_de = "dummy"
        previous_description_fr = "dummy"

        # Create an Excel file
        # file_path = "basic_excel.xlsx"
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            current_row = 1  # Starting row for writing data
            worksheet = (
                writer.sheets["Sheet1"]
                if "Sheet1" in writer.sheets
                else writer.book.create_sheet("Sheet1")
            )
            data = []

            # Title
            schema_version = yaml_data.get("version", "")
            model = yaml_data.get("model", {})
            revision = model.get("revision", "")
            revision_date = model.get("revision_date", "")

            title_font = Font(size=20, bold=True, name="Arial")

            cell = worksheet["A1"]
            cell.value = f"Geology Data Model"
            cell.font = title_font
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
            worksheet.row_dimensions[current_row].height = 30

            cell = worksheet["A2"]
            cell.value = "Revision"
            cell.font = Font(size=12, bold=True, name="Arial")
            cell = worksheet["B2"]
            cell.value = revision
            cell.font = Font(size=12, bold=True, name="Arial")

            cell = worksheet["A3"]
            cell.value = "Date"
            cell.font = Font(size=12, bold=True, name="Arial")
            cell = worksheet["B3"]
            cell.value = revision_date
            cell.font = Font(size=12, bold=True, name="Arial")

            cell = worksheet["A4"]
            cell.value = "Schema"
            cell.font = Font(size=12, bold=True, name="Arial")
            cell = worksheet["B4"]
            cell.value = schema_version
            cell.font = Font(size=12, bold=True, name="Arial")

            # Format the headers (bold, size 14, center alignment)
            current_row += 4

            worksheet[f"C{current_row}"] = "Deutsch"
            worksheet[f"D{current_row}"] = "Français"
            worksheet[f"E{current_row}"] = "Type"
            worksheet[f"F{current_row}"] = "Mandatory"
            worksheet[f"G{current_row}"] = "Cardinality"

            header_font = Font(size=14, bold=True, color="dddddd", name="Arial")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_fill = PatternFill(
                start_color="333333", end_color="dddddd", fill_type="solid"
            )

            # Apply the formatting to the headers
            for col in [chr(i) for i in range(ord("A"), ord("G") + 1)]:
                cell = worksheet[f"{col}{current_row}"]
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill

            current_row += 1

            for theme in yaml_data.get("themes", []):
                theme_name = theme.get("name", "")
                # theme_desc_de = theme.get("description", {}).get("de", "")
                # theme_desc_fr = theme.get("description", {}).get("fr", "")
                for cls in theme.get("classes", []):
                    data = []
                    class_name = cls.get("name", "")
                    if "Unconsolidated_Deposits_PLG" in class_name:
                        logger.debug(cls.get("description", {}))
                    abrev = cls.get("abrev", "")
                    table = cls.get("table", "")
                    description_de = cls.get("description", {}).get("de", "")
                    description_fr = cls.get("description", {}).get("fr", "")
                    logger.debug(
                        description_de.strip()
                        if description_de != previous_description_de
                        else ""
                    )

                    # Iterate over attributes
                    for attr in cls.get("attributes", []):
                        attr_name = attr.get("name", "")
                        attr_type = attr.get("att_type", "")
                        attr_value = attr.get("value", "")
                        attr_desc_de = attr.get("description", {}).get("de", "")
                        attr_desc_fr = attr.get("description", {}).get("fr", "")
                        cardinality = attr.get("cardinality", "")
                        mandatory = attr.get("mandatory", False)

                        (
                            theme_name if theme_name != previous_theme else "",
                        )  # Omit repeating theme name
                        (
                            class_name if class_name != previous_class else "",
                        )  # Omit repeating class name
                        (
                            description_de
                            if description_de != previous_description_de
                            else "",
                        )
                        (
                            description_fr
                            if description_fr != previous_description_fr
                            else "",
                        )

                        # Prepare the row data
                        row_data = [
                            # "Abbreviation": abrev if abrev != previous_abrev else "",  # Omit repeating class name
                            # "Table Name": table if table != previous_table else "",  # Omit repeating class name,
                            attr_name,
                            attr_desc_de,
                            attr_desc_fr,
                            attr_type,
                            attr_value,
                            "yes" if mandatory else "no",
                            f"[{cardinality}]",
                        ]
                        data.append(row_data)

                    # Theme
                    if theme_name != previous_theme:
                            cell = worksheet.cell(
                                row=current_row, column=1, value="Theme"
                            )
                            cell.font = Font(bold=True, name="Arial")
                            cell = worksheet.cell(
                                row=current_row, column=2, value=theme_name
                            )
                            cell.font = font = Font(size=20, bold=True, name="Arial")

                    # Classe
                    if class_name != previous_class:
                            cell = worksheet.cell(
                                row=current_row + 1, column=1, value="Class"
                            )
                            cell.font = Font(bold=True, name="Arial")

                            cell = worksheet.cell(
                                row=current_row + 1, column=2, value=class_name
                            )
                            cell.font = Font(size=14, bold=True, name="Arial")

                            my_align = Alignment(wrapText=True, vertical="top")
                            for col_range in range(1, 12):
                                cell_title = worksheet.cell(current_row + 1, col_range)
                                cell_title.fill = PatternFill(
                                    start_color="dddddd",
                                    end_color="dddddd",
                                    fill_type="solid",
                                )
                            worksheet.cell(
                                row=current_row + 2, column=3, value=description_de
                            )
                            worksheet.cell(
                                row=current_row + 2, column=4, value=description_fr
                            )

                            worksheet.cell(
                                row=current_row + 2, column=3
                            ).alignment = my_align
                            worksheet.cell(
                                row=current_row + 2, column=4
                            ).alignment = my_align
                            cell = worksheet.cell(
                                row=current_row + 2, column=1, value="Description"
                            )
                            cell.font = Font(bold=True, name="Arial")
                            cell.alignment = my_align

                            worksheet.cell(row=current_row + 3, column=2, value=abrev)
                            cell = worksheet.cell(
                                row=current_row + 3, column=1, value="Abrev"
                            )
                            cell.font = Font(bold=True, name="Arial")

                            worksheet.cell(row=current_row + 4, column=2, value=table)
                            cell = worksheet.cell(
                                row=current_row + 4, column=1, value="Table"
                            )
                            cell.font = Font(bold=True, name="Arial")

                            cell = worksheet.cell(
                                row=current_row + 5, column=1, value="Attributes"
                            )
                            cell.font = Font(bold=True, name="Arial")

                            worksheet.row_dimensions[current_row].height = 30
                            worksheet.row_dimensions[current_row + 1].height = 20

                            worksheet.row_dimensions[current_row + 2].height = 100

                            column_index = 0

                            current_row += 5

                            for attr in data:
                                (
                                    attr_name,
                                    attr_desc_de,
                                    attr_desc_fr,
                                    attr_type,
                                    attr_value,
                                    mandatory,
                                    cardinality,
                                ) = attr

                                worksheet.cell(
                                    row=current_row, column=2, value=attr_name.upper()
                                )  # Indented description
                                worksheet.cell(
                                    row=current_row, column=3, value=attr_desc_de
                                ).alignment = my_align  # Indented cardinality
                                worksheet.cell(
                                    row=current_row, column=4, value=attr_desc_fr
                                ).alignment = my_align  # Indented mandatory

                                worksheet.cell(
                                    row=current_row, column=5, value=attr_type
                                )
                                worksheet.cell(
                                    row=current_row, column=6, value=mandatory
                                )
                                worksheet.cell(
                                    row=current_row, column=7, value=cardinality
                                )
                                current_row += 1  # Move down for each attribute

                            current_row += 2  # Next class

                            dim_holder = DimensionHolder(worksheet=worksheet)

                            for col in range(
                                worksheet.min_column, worksheet.max_column + 1
                            ):
                                column_name = get_column_letter(col)
                                width = 20
                                if column_name.upper() in ["C", "D"]:
                                    width = 75
                                if column_name.upper() in ["E", "F", "G"]:
                                    width = 20

                                dim_holder[column_name] = ColumnDimension(
                                    worksheet, min=col, max=col, width=width
                                )
                                worksheet.column_dimensions = dim_holder

                    # Update previous_theme and previous_class
                    data = []
                    previous_theme = theme_name
                    previous_class = class_name
                    previous_abrev = abrev
                    previous_table = table
                    previous_description_de = description_de
                    previous_description_fr = description_fr

    def export_to_yaml(self, file_path):
        """Export the yaml_data structure to a YAML file."""
        output_data = {
            key: dict(value) if isinstance(value, OrderedDict) else value
            for key, value in self.yaml_data.items()
        }
        with open(file_path, "w", encoding="utf-8") as yaml_file:
            yaml.dump(output_data, yaml_file, allow_unicode=True, sort_keys=False)

    def import_from_yaml(self, file_path):
        """Import data from a YAML file into the yaml_data structure."""
        with open(file_path, "r", encoding="utf-8") as yaml_file:
            self.yaml_data = yaml.load(yaml_file, Loader=yaml.FullLoader)


if __name__ == "__main__":
    # Example Usage
    xlsx_file = "modelclass.xlsx"
    datamodel = Datamodel()
    datamodel.import_from_yaml("datamodel.yaml")
    logger.info(datamodel.yaml_data)

    logger.info("Export to Excel")
    # datamodel.export_to_yaml("output.yaml")
    datamodel.export_to_excel(xlsx_file)
    logger.info("Import from  Excel")
    datamodel.import_from_excel(xlsx_file)

    logger.info("Export to YAML")

    datamodel.export_to_yaml("modelclass.yaml")
