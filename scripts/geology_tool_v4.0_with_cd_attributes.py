#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Geology Mapping Tool Generator - Version 4.0
Intégration automatique des attributs CD depuis datamodel.yaml
"""

import json
import os

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from rich.console import Console

console = Console()

name_mapping = {
    "Unconsolidated_Deposits_PT": "Deposits_PT",
    "Instability_Structures_PT": "Instability_PT",
    "Instability_Structures_L": "Instability_L",
    "Instabilities_PLG": "Instability_PLG",
    "Glacial_Structures_PT": "Glacial_PT",
    "Glacial_and_Periglacial_Structu": "Glacial/Periglacial",
    "Glacial_Structures_PLG": "Glacial_PLG",
    "Erosional_Structures_PT": "Erosion_PT",
    "Erosional_Structures_L": "Erosion_L",
    "Karstic_Structures_PT": "Karst_PT",
    "Karstic_Structures_PLG": "Karst_PLG",
    "Alluvial_and_Lacustrine_Structu": "Alluvial/Lacustrine",
    "Deformation_Structures_PT": "Deformation_PT",
    "Deformation_Structures_L": "Deformation_L",
    "Deformation_Structures_PLG": "Deformation_PLG",
    "Tectonic_Boundaries_L": "Tectonic_L",
    "Folds_PT": "Folds_PT",
    "Lineation_PT": "Lineation_PT",
    "Planar_Structures_PT": "Planar_PT",
    "Anomalies_PT": "Anomaly_PT",
    "Indication_of_Resources_PT": "Resources_PT",
    "Mineralised_Zone_L": "Mineral_L",
    "Sedimentary_Structures_PT": "Sediment_PT",
    "Type_Localities_PT": "Locality_PT",
    "Prominent_Lithological_Features": "Lithology",
    "Miscellaneous_PT": "Misc_PT",
    "Geological_Outlines_L": "Geology_L",
    "Slope_Bedrock_PT": "Slope_PT",
    "Contour_Lines_Bedrock_L": "Contour_Bedrock_L",
    "Modelled_Water_Table_PT": "WaterTable_PT",
    "Contour_Lines_Hydro_L": "Contour_Hydro_L",
    "Archaeology_PT": "Arch_PT",
    "Archaeology_L": "Arch_L",
    "Archaeology_PLG": "Arch_PLG",
    "Exploitation_Geomaterials_PT": "Geomaterials_PT",
    "Exploitation_Geomaterials_L": "Geomaterials_L",
    "Exploitation_Geomaterials_PLG": "Geomaterials_PLG",
    "Boreholes_PT": "Boreholes_PT",
    "Artificial_Surface_Modification": "Surface_Mod",
    "Construction_PT": "Construct_PT",
    "Construction_L": "Construct_L",
    "Palaeohydrology_L": "PalaeoHydro_L",
    "Subsurface_Water_L": "SubWater_L",
    "Surface_Water_PT": "SurfWater_PT",
    "Surface_Water_L": "SurfWater_L",
    "Surface_Water_PLG": "SurfWater_PLG",
}


console.print("=" * 70)
console.print("  🪨 GEOLOGY MAPPING TOOL GENERATOR v4.0")
console.print("  📋 Intégration automatique des attributs CD")
console.print("=" * 70)
console.print("\n📊 Chargement des données du projet...")

PROJECT_PATH = "../exports/2025-10-24"


def load_json(filename):
    filepath = os.path.join(PROJECT_PATH, filename)
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
            console.print(
                f"   ✅ {filename}: {len(data) if isinstance(data, (dict, list)) else 'OK'}"
            )
            return data
    except FileNotFoundError:
        console.print(f"   ⚠️  {filename}: NON TROUVÉ")
        return {} if "Att" not in filename else []
    except Exception as e:
        console.print(f"   ❌ {filename}: ERREUR - {e}")
        return {} if "Att" not in filename else []


# Load datamodel.yaml
console.print("\n📋 Chargement du datamodel.yaml...")
datamodel_path = "datamodel.yaml"
class_cd_attributes = {}
PROJECT_PATH = None

try:
    with open(datamodel_path, "r", encoding="utf-8") as f:
        datamodel = yaml.safe_load(f)

    model = datamodel.get("model")
    if model:
        PROJECT_PATH = model.get("sources_dir")
        if not os.path.isdir(PROJECT_PATH):
            console.print(f"Non existing {PROJECT_PATH}. Exiting")

            sys.exit(2)

    # Extract CD attributes for each class
    for theme in datamodel.get("themes", []):
        for cls in theme.get("classes", []):
            class_name = cls.get("name")
            cd_attrs = []

            for attr in cls.get("attributes", []):
                name = attr.get("name")
                if name == "kind":
                    val = attr.get("value")
                    # console.print(f"('{val}','{class_name}'),")
                if attr.get("att_type") == "CD":
                    cd_domain = attr.get("value", "")
                    # Ignore dummy/N/A domains
                    if cd_domain and cd_domain not in ["N/A", "dummy", "DUMMY"]:
                        domain_dict = {
                            "name": attr.get("name"),
                            "cd_domain": cd_domain,
                            "desc_de": attr.get("description", {}).get("de", ""),
                            "desc_fr": attr.get("description", {}).get("fr", ""),
                            "mandatory": attr.get("mandatory", False),
                        }
                        cd_attrs.append(domain_dict)
                        # console.print(domain_dict)

            if cd_attrs:
                class_cd_attributes[class_name] = cd_attrs

    console.print(
        f"   ✅ Datamodel chargé: {len(class_cd_attributes)} classes avec attributs CD"
    )

except Exception as e:
    console.print(f"   ❌ Erreur lors du chargement du datamodel: {e}")
    import traceback

    traceback.print_exc()

# Load main data
geol_mapping_units = load_json("Geol_Mapping_Unit.json")
geol_mapping_unit_att = load_json("Geol_Mapping_Unit_Att.json")
litho = load_json("Litho.json")
chrono = load_json("Chrono.json")
tecto = load_json("Tecto.json")
litstrat_formation_bank = load_json("Litstrat_Formation_Bank.json")
litstrat_unco = load_json("Litstrat_Unco.json")
system_data = load_json("System.json")

# Load gcoverp_export_simple.json
print("\n📋 Chargement des coded domains et subtypes...")
coded_domains = {}
subtypes = {}
gcoverp_file = os.path.join(PROJECT_PATH, "gcoverp_export_simple.json")

try:
    with open(gcoverp_file, "r", encoding="utf-8") as f:
        gcoverp_data = json.load(f)

    if "subtypes" in gcoverp_data:
        subtypes = gcoverp_data["subtypes"]
        console.print(f"   ✅ Subtypes chargés: {len(subtypes)} entrées")

    if "coded_domain" in gcoverp_data:
        for domain_name, domain_data in gcoverp_data["coded_domain"].items():
            if isinstance(domain_data, dict) and "codedValues" in domain_data:
                coded_domains[domain_name] = domain_data["codedValues"]
        console.print(f"   ✅ Coded domains chargés: {len(coded_domains)} domaines")

except FileNotFoundError:
    console.print(f"   ⚠️  gcoverp_export_simple.json NON TROUVÉ")
except Exception as e:
    console.print(f"   ❌ Erreur lors du chargement: {e}")
    import traceback

    traceback.print_exc()


wb = Workbook()

# Styling
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
locked_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
editable_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# =================================================================
# HELPER FUNCTIONS
# =================================================================


def create_reference_sheet(wb, sheet_name, data_dict):
    ws = wb.create_sheet(sheet_name)
    ws["A1"], ws["B1"], ws["C1"] = "Code", "Label", "Display"
    for col in ["A1", "B1", "C1"]:
        ws[col].fill = header_fill
        ws[col].font = header_font

    # Sort by Label ascending, but push codes < 1e6 to the end
    items = sorted(data_dict.items(), key=lambda x: (int(x[0]) < 1_000_000, str(x[1])))

    row = 2
    for code, label in items:
        ws[f"A{row}"] = int(code)
        ws[f"B{row}"] = str(label)
        ws[f"C{row}"] = f"{code}: {label}"
        row += 1

    return row - 1


def add_dropdown(ws, col_range, ref_sheet, ref_range, prompt_text, title):
    dv = DataValidation(
        type="list", formula1=f"={ref_sheet}!${ref_range}", allow_blank=True
    )
    dv.prompt = prompt_text
    dv.promptTitle = title
    ws.add_data_validation(dv)
    dv.add(col_range)


def extract_subtypes_by_prefix(subtypes_dict, prefix):
    """Extract subtypes that match a given prefix (flat dictionary)"""
    result = {}
    for code, label in subtypes_dict.items():
        if str(code).startswith(str(prefix)):
            result[code] = label
    return result


# =================================================================
# REFERENCE SHEETS
# =================================================================
print("\n📚 Création des feuilles de référence...")

gmu_last_row = create_reference_sheet(wb, "Ref_GeolMappingUnit", geol_mapping_units)
litho_last_row = create_reference_sheet(wb, "Ref_Litho", litho)
chrono_last_row = create_reference_sheet(wb, "Ref_Chrono", chrono)
tecto_last_row = create_reference_sheet(wb, "Ref_Tecto", tecto)
litstrat_fm_last_row = create_reference_sheet(
    wb, "Ref_LithoStratFormation", litstrat_formation_bank
)
litstrat_unco_last_row = create_reference_sheet(wb, "Ref_LithoStratUnco", litstrat_unco)
system_last_row = create_reference_sheet(wb, "Ref_System", system_data)

# GMU Attributes
ws_ref_gmu_att = wb.create_sheet("Ref_GMU_Attributes")
headers = [
    "GEOL_MAPPING_UNIT",
    "LITSTRAT_FORMATION_BANK",
    "CHRONO_TOP",
    "CHRONO_BASE",
    "LITHO_MAIN",
    "LITHO_SEC",
    "LITHO_TER",
]
for idx, header in enumerate(headers, start=1):
    cell = ws_ref_gmu_att.cell(1, idx, header)
    cell.fill = header_fill
    cell.font = header_font

row = 2
for att in geol_mapping_unit_att:
    ws_ref_gmu_att[f"A{row}"] = att["GEOL_MAPPING_UNIT"]
    ws_ref_gmu_att[f"B{row}"] = att["LITSTRAT_FORMATION_BANK"]
    ws_ref_gmu_att[f"C{row}"] = att["CHRONO_TOP"]
    ws_ref_gmu_att[f"D{row}"] = att["CHRONO_BASE"]
    ws_ref_gmu_att[f"E{row}"] = att["LITHO_MAIN"]
    ws_ref_gmu_att[f"F{row}"] = att["LITHO_SEC"] if att["LITHO_SEC"] != 0 else ""
    ws_ref_gmu_att[f"G{row}"] = att["LITHO_TER"] if att["LITHO_TER"] != 0 else ""
    row += 1

# Create reference sheets for all coded domains
console.print("\n📋 Création des feuilles pour coded domains...")
cd_refs = {}

for cd_name, cd_values in coded_domains.items():
    if cd_values:  # Only create if not empty
        ref_name = f"Ref_CD_{cd_name}"[:31]  # Truncate to 31 chars
        try:
            last_row = create_reference_sheet(wb, ref_name, cd_values)
            cd_refs[cd_name] = {"sheet_name": ref_name, "last_row": last_row}
        except Exception as e:
            console.print(f"   ⚠️  Impossible de créer {ref_name}: {e}")

console.print(f"   ✅ {len(cd_refs)} feuilles de coded domains créées")

# Extract subtypes by prefix
console.print("\n📋 Extraction des subtypes par classe...")

# Map class prefixes to their data

classes_config = [
    ("12901", "Fossils_PT"),
    ("14401", "Unconsolidated_Deposits_PT"),
    ("11601", "Instability_Structures_PT"),
    ("11701", "Instability_Structures_L"),
    ("15801", "Instabilities_PLG"),
    ("11201", "Glacial_Structures_PT"),
    ("11301", "Glacial_and_Periglacial_Structures_L"),
    ("11401", "Glacial_Structures_PLG"),
    ("11001", "Erosional_Structures_PT"),
    ("11101", "Erosional_Structures_L"),
    ("11901", "Karstic_Structures_PT"),
    ("12001", "Karstic_Structures_PLG"),
    ("10901", "Alluvial_and_Lacustrine_Structures_L"),
    ("14601", "Deformation_Structures_PT"),
    ("14701", "Deformation_Structures_L"),
    ("14801", "Deformation_Structures_PLG"),
    ("14901", "Tectonic_Boundaries_L"),
    ("13601", "Folds_PT"),
    ("13701", "Lineation_PT"),
    ("13801", "Planar_Structures_PT"),
    ("12801", "Anomalies_PT"),
    ("13201", "Indication_of_Resources_PT"),
    ("13301", "Mineralised_Zone_L"),
    ("13401", "Sedimentary_Structures_PT"),
    ("13501", "Type_Localities_PT"),
    ("13101", "Prominent_Lithological_Features_L"),
    ("15501", "Miscellaneous_PT"),
    ("13001", "Geological_Outlines_L"),
    ("14201", "Slope_Bedrock_PT"),
    ("13901", "Contour_Lines_Bedrock_L"),
    ("14101", "Modelled_Water_Table_PT"),
    ("14001", "Contour_Lines_Hydro_L"),
    ("10101", "Archaeology_PT"),
    ("10201", "Archaeology_L"),
    ("10301", "Archaeology_PLG"),
    ("10601", "Exploitation_Geomaterials_PT"),
    ("10701", "Exploitation_Geomaterials_L"),
    ("10801", "Exploitation_Geomaterials_PLG"),
    ("10501", "Boreholes_PT"),
    ("10401", "Artificial_Surface_Modifications_PLG"),
    ("12101", "Construction_PT"),
    ("12201", "Construction_L"),
    ("12301", "Palaeohydrology_L"),
    ("12401", "Subsurface_Water_L"),
    ("12501", "Surface_Water_PT"),
    ("12601", "Surface_Water_L"),
    ("12701", "Surface_Water_PLG"),
]

# Create reference sheets for each class's KIND values
subtype_refs = {}
for prefix, class_name in classes_config:
    kind_values = extract_subtypes_by_prefix(subtypes, prefix)
    if kind_values:
        ref_name = f"Ref_KIND_{class_name}"[:31]
        subtype_refs[class_name] = create_reference_sheet(wb, ref_name, kind_values)
        console.print(f"   ✅ {ref_name}: {len(kind_values)} KINDs")

console.print(f"\n✅ Total: {len(subtype_refs)} feuilles de subtypes créées")

# =================================================================
# MAIN SHEETS (Bedrock, Unconsolidated, Fossils)
# =================================================================
console.print("\n🪨 Création des feuilles principales...")

# BEDROCK_PLG
ws_bedrock = wb.create_sheet("Bedrock_PLG", 0)
headers_bedrock = [
    "User_Code",
    "GEOL_MAPPING_UNIT_Display",
    "GEOL_MAPPING_UNIT",
    "TECTO_Display",
    "TECTO",
    "LITSTRAT_FORMATION_BANK_Display",
    "CHRONO_TOP_Display",
    "CHRONO_BASE_Display",
    "LITHO_MAIN_Display",
    "LITHO_SEC_Display",
    "LITHO_TER_Display",
]
for col, header in enumerate(headers_bedrock, start=1):
    cell = ws_bedrock.cell(1, col, header)
    cell.fill = header_fill
    cell.font = header_font

widths = [15, 50, 20, 40, 15, 50, 35, 35, 35, 35, 35]
for i, width in enumerate(widths, start=1):
    ws_bedrock.column_dimensions[chr(64 + i)].width = width

# Hide
for col in ["C", "E"]:
    ws_bedrock.column_dimensions[col].hidden = True

add_dropdown(
    ws_bedrock,
    "B2:B1000",
    "Ref_GeolMappingUnit",
    f"C$2:C${gmu_last_row}",
    "Sélectionnez une unité",
    "GEOL_MAPPING_UNIT",
)
add_dropdown(
    ws_bedrock,
    "D2:D1000",
    "Ref_Tecto",
    f"C$2:C${tecto_last_row}",
    "Sélectionnez une unité tectonique",
    "TECTO",
)

for row_num in range(2, 1001):
    ws_bedrock[f"A{row_num}"].fill = editable_fill
    ws_bedrock[f"B{row_num}"].fill = editable_fill
    ws_bedrock[f"C{row_num}"] = (
        f'=IF(B{row_num}="","",VALUE(LEFT(B{row_num},FIND(":",B{row_num})-1)))'
    )
    ws_bedrock[f"D{row_num}"].fill = editable_fill
    ws_bedrock[f"E{row_num}"] = (
        f'=IF(D{row_num}="","",VALUE(LEFT(D{row_num},FIND(":",D{row_num})-1)))'
    )

    lookup_cols = [
        ("F", "B", "Ref_LithoStratFormation"),
        ("G", "C", "Ref_Chrono"),
        ("H", "D", "Ref_Chrono"),
        ("I", "E", "Ref_Litho"),
        ("J", "F", "Ref_Litho"),
        ("K", "G", "Ref_Litho"),
    ]

    for target_col, source_col, ref_sheet in lookup_cols:
        ws_bedrock[f"{target_col}{row_num}"] = (
            f'=IF(C{row_num}="","",IFERROR('
            f'INDEX({ref_sheet}!$C:$C,MATCH(VLOOKUP(C{row_num},Ref_GMU_Attributes!$A:${source_col},COLUMN({source_col}:${source_col}),FALSE),{ref_sheet}!$A:$A,0)),""))'
        )
        ws_bedrock[f"{target_col}{row_num}"].fill = locked_fill

for idx, code in enumerate(["BEDROCK_001", "BEDROCK_002"], start=2):
    ws_bedrock[f"A{idx}"] = code

# UNCONSOLIDATED_DEPOSITS_PLG
ws_unco = wb.create_sheet("Unconsolidated_Deposits_PLG", 1)
headers_unco = [
    "User_Code",
    "LITSTRAT_Display",
    "LITSTRAT",
    "LITHO_Display",
    "LITHO",
    "CHRONO_TOP_Display",
    "CHRONO_TOP",
    "CHRONO_BASE_Display",
    "CHRONO_BASE",
]
for col, header in enumerate(headers_unco, start=1):
    cell = ws_unco.cell(1, col, header)
    cell.fill = header_fill
    cell.font = header_font

for col in ["C", "E", "G", "I"]:
    ws_unco.column_dimensions[col].hidden = True

for i in range(1, 10):
    ws_unco.column_dimensions[chr(64 + i)].width = 15 if i in [1, 3, 5, 7, 9] else 50

add_dropdown(
    ws_unco,
    "B2:B1000",
    "Ref_LithoStratUnco",
    f"C$2:C${litstrat_unco_last_row}",
    "Sélectionnez une unité",
    "LITSTRAT",
)
add_dropdown(
    ws_unco,
    "D2:D1000",
    "Ref_Litho",
    f"C$2:C${litho_last_row}",
    "Sélectionnez une lithologie",
    "LITHO",
)
add_dropdown(
    ws_unco,
    "F2:F1000",
    "Ref_Chrono",
    f"C$2:C${chrono_last_row}",
    "Chronostratigraphie (top)",
    "CHRONO_TOP",
)
add_dropdown(
    ws_unco,
    "H2:H1000",
    "Ref_Chrono",
    f"C$2:C${chrono_last_row}",
    "Chronostratigraphie (base)",
    "CHRONO_BASE",
)

for row_num in range(2, 1001):
    for col in ["A", "B", "D", "F", "H"]:
        ws_unco[f"{col}{row_num}"].fill = editable_fill

    ws_unco[f"C{row_num}"] = (
        f'=IF(B{row_num}="","",VALUE(LEFT(B{row_num},FIND(":",B{row_num})-1)))'
    )
    ws_unco[f"E{row_num}"] = (
        f'=IF(D{row_num}="","",VALUE(LEFT(D{row_num},FIND(":",D{row_num})-1)))'
    )
    ws_unco[f"G{row_num}"] = (
        f'=IF(F{row_num}="","",VALUE(LEFT(F{row_num},FIND(":",F{row_num})-1)))'
    )
    ws_unco[f"I{row_num}"] = (
        f'=IF(H{row_num}="","",VALUE(LEFT(H{row_num},FIND(":",H{row_num})-1)))'
    )

for idx, code in enumerate(["UNCO_001", "UNCO_002"], start=2):
    ws_unco[f"A{idx}"] = code


console.print("   ✅ Bedrock, Unconsolidated créés")

# =================================================================
# SEPARATE SHEETS FOR OTHER CLASSES WITH CD ATTRIBUTES
# =================================================================
console.print("\n📏 Création des feuilles pour classes avec attributs CD...")

sheet_idx = 2  # no more fossils
for prefix, class_name in classes_config:
    sheet_idx += 1

    # Truncate sheet name if > 31 characters
    sheet_name = class_name[:31] if len(class_name) > 31 else class_name
    # Shorten sheet name
    sheet_name = name_mapping.get(class_name, class_name)

    console.print(f"   ✅ {sheet_name} (class_name)")

    ws = wb.create_sheet(sheet_name, sheet_idx)

    # Build headers: User_Code, KIND_Display, KIND, then CD attributes
    headers = ["User_Code", "KIND_Display", "KIND"]

    # Get CD attributes for this class from datamodel
    cd_attrs = class_cd_attributes.get(class_name, [])

    # Add CD attribute columns
    for cd_attr in cd_attrs:
        attr_name = cd_attr["name"].upper()
        headers.append(f"{attr_name}_Display")
        headers.append(attr_name)

    headers.append("Notes")

    # Set headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font

    # Set column widths
    for i in range(1, len(headers) + 1):
        head_name = None
        try:
            head_name = headers[i - 1]
            console.print(f"     {chr(64 + i)} - {head_name}")
        except Exception as e:
            console.print(e)
        if head_name and (
            head_name in ("Notes", "User_Code") or head_name.endswith("_Display")
        ):
            ws.column_dimensions[chr(64 + i)].width = 20 if i <= 3 else 35
        else:
            ws.column_dimensions[chr(64 + i)].hidden = True
            # ws.column_dimensions[chr(64 + i)].width = 9

    # Add KIND dropdown if available
    if class_name in subtype_refs:
        ref_name = f"Ref_KIND_{class_name}"[:31]
        add_dropdown(
            ws,
            "B2:B1000",
            ref_name,
            f"C$2:C${subtype_refs[class_name]}",
            f"Sélectionnez un KIND",
            "KIND",
        )

    # Add dropdowns for CD attributes
    col_idx = 4  # Start after User_Code, KIND_Display, KIND
    for cd_attr in cd_attrs:
        cd_domain = cd_attr["cd_domain"]
        attr_name = cd_attr["name"].upper()

        if cd_domain in cd_refs:
            ref_info = cd_refs[cd_domain]
            col_letter_display = chr(64 + col_idx)

            add_dropdown(
                ws,
                f"{col_letter_display}2:{col_letter_display}1000",
                ref_info["sheet_name"],
                f"C$2:C${ref_info['last_row']}",
                cd_attr["desc_de"],
                attr_name,
            )

        col_idx += 2  # Skip _Display and code columns

    # Add formulas and styling
    for row_num in range(2, 1001):  # TODO is enough
        # User_Code and KIND_Display editable
        ws[f"A{row_num}"].fill = editable_fill
        ws[f"B{row_num}"].fill = editable_fill

        # Extract KIND code
        ws[f"C{row_num}"] = (
            f'=IF(B{row_num}="","",VALUE(LEFT(B{row_num},FIND(":",B{row_num})-1)))'
        )

        # Handle CD attribute columns
        col_idx = 4
        for cd_attr in cd_attrs:
            col_letter_display = chr(64 + col_idx)
            col_letter_code = chr(64 + col_idx + 1)

            # Display column is editable
            ws[f"{col_letter_display}{row_num}"].fill = editable_fill

            # Code column has formula
            ws[f"{col_letter_code}{row_num}"] = (
                f'=IF({col_letter_display}{row_num}="","",VALUE(LEFT({col_letter_display}{row_num},FIND(":",{col_letter_display}{row_num})-1)))'
            )

            col_idx += 2

        # Notes column
        ws[f"{chr(64 + len(headers))}{row_num}"].fill = editable_fill

    # Add example code
    ws["A2"] = f"{class_name.upper().replace('_', '')[:10]}_001"

    cd_count = len(cd_attrs)
    console.print(f"     CD attributs: {cd_count}")

# =================================================================
# FINALIZATION
# =================================================================

if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

output_path = "outputs/geology_mapping_tool_v4.0.xlsx"
wb.save(output_path)

console.print("\n" + "=" * 70)
console.print("  ✅ VERSION 4.0 - AVEC ATTRIBUTS CD!")
console.print("=" * 70)
console.print(f"\n📚 Fichier: {output_path}")
console.print(f"\n📊 Structure:")
console.print(f"   - 2 feuilles principales (Bedrock et Unconsolidated)")
console.print(f"   - {len(classes_config)} feuilles pour classes linéaires avec attributs CD")
console.print(f"   - {len(subtype_refs)} feuilles de référence KIND")
console.print(f"   - {len(cd_refs)} feuilles de coded domains")
console.print(f"   - Total: {len(wb.sheetnames)} feuilles")
console.print("\n" + "=" * 70)
