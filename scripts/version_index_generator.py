#!/usr/bin/env python3
"""
Generate an HTML index page for available datamodel versions in the S3 bucket.

Versioning convention
---------------------
  x.y   = schema version  (what users / geologists care about)
  x.y.z = tooling/patch   (what gets physically stored in S3)

The index groups all x.y.z patches under their x.y parent and renders
files from the latest patch — conceptually a symlink x.y → newest x.y.z.

Translations are loaded from  <repo_root>/translation.xlsx  (sheet
"translations", columns: msg_id | de | fr | it | en).

Usage
-----
  python scripts/version_index_generator.py [--lang fr] [--no-upload]

Environment variables
---------------------
  S3_BUCKET          required
  AWS_PROFILE        optional  (ignored in CI)
  CLOUDFRONT_DOMAIN  default: dubious.cloud  (set to "" or "none" to use S3 URL)
  S3_PREFIX          default: datamodel/
"""

import os
import sys
import re
from collections import OrderedDict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

import boto3
import click
from jinja2 import Environment, FileSystemLoader, select_autoescape

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent          # …/scripts/
REPO_ROOT   = SCRIPT_DIR.parent                       # …/ (project root)
TEMPLATE_DIR = SCRIPT_DIR / "templates"
TRANSLATION_FILE = REPO_ROOT / "translations.xlsx"


# ---------------------------------------------------------------------------
# Translations
# ---------------------------------------------------------------------------

def load_translations(xlsx_path: Path, lang: str) -> Dict[str, str]:
    """
    Read translation.xlsx and return a dict {msg_id: translated_string}
    for the requested language.  Falls back to 'de' when a cell is empty.

    Expected sheet layout:
        msg_id | de | fr | it | en
    """
    try:
        import openpyxl
    except ImportError:
        print("❌  openpyxl is required to read translation.xlsx  →  pip install openpyxl")
        sys.exit(1)

    if not xlsx_path.exists():
        print(f"❌  Translation file not found: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Build column index from header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        col_msgid = headers.index("msg_id")
        col_lang  = headers.index(lang)
        col_de    = headers.index("de")     # fallback
    except ValueError as exc:
        print(f"❌  Missing column in translation.xlsx: {exc}")
        sys.exit(1)

    translations: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        msg_id = row[col_msgid]
        if not msg_id:
            continue
        value = row[col_lang] or row[col_de] or f"[{msg_id}]"
        translations[str(msg_id)] = str(value)

    wb.close()
    return translations


# ---------------------------------------------------------------------------
# S3 helpers
# ---------------------------------------------------------------------------

def get_s3_client(profile_name: str = None):
    session = (
        boto3.Session(profile_name=profile_name) if profile_name else boto3.Session()
    )
    return session.client("s3")


def parse_version(version_str: str) -> tuple:
    """Parse version string into a sortable tuple."""
    parts = re.findall(r"\d+|\D+", version_str)
    return tuple(
        (0, int(p)) if p.isdigit() else (1, p)
        for p in parts
    )


def list_versions(s3_client, bucket_name: str, prefix: str) -> List[str]:
    """Return all version directories (newest first)."""
    versions = []
    try:
        paginator = s3_client.get_paginator("list_objects_v2")
        for page in paginator.paginate(Bucket=bucket_name, Prefix=prefix, Delimiter="/"):
            for p in page.get("CommonPrefixes", []):
                v = p["Prefix"].replace(prefix, "").rstrip("/")
                if v:
                    versions.append(v)
        versions.sort(key=parse_version, reverse=True)
    except Exception as exc:
        print(f"Error listing versions: {exc}")
        sys.exit(1)
    return versions


def group_versions_by_minor(versions: List[str]) -> "OrderedDict[str, List[str]]":
    """
    Group x.y.z patch versions under their x.y schema-version key.

    Input  (newest first): ["1.3.1", "1.3.0", "1.2.4", "1.2.3"]
    Output:
        OrderedDict {
            "1.3": ["1.3.1", "1.3.0"],
            "1.2": ["1.2.4", "1.2.3"],
        }
    Legacy x.y-only entries in S3 are kept as their own group.
    """
    groups: OrderedDict = OrderedDict()
    for v in versions:
        parts = v.split(".")
        key = ".".join(parts[:2]) if len(parts) >= 3 else v
        groups.setdefault(key, []).append(v)
    return groups


EXCLUDED_EXTENSIONS = {".css", ".map"}
FILE_ICONS = {
    "pdf": "📄", "docx": "📝", "html": "🌐",
    "odt": "📄", "md": "📋", "json": "🔧",
    "yaml": "⚙️", "yml": "⚙️",
}


def _icon(filename: str) -> str:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    return FILE_ICONS.get(ext, "📎")


def _fmt_size(size_bytes: int) -> str:
    if size_bytes == 0:
        return "0 B"
    names = ["B", "KB", "MB", "GB"]
    i = 0
    while size_bytes >= 1024 and i < len(names) - 1:
        size_bytes /= 1024.0
        i += 1
    return f"{size_bytes:.1f} {names[i]}"


def get_files_for_version(
    s3_client, bucket_name: str, version: str, prefix: str
) -> Dict[str, List[Dict[str, Any]]]:
    """Return files for *version*, organised by language key."""
    version_prefix = f"{prefix}{version}/"
    files_by_lang: Dict[str, List] = {k: [] for k in ("root", "de", "fr", "it", "en")}

    try:
        paginator = s3_client.get_paginator("list_objects_v2")
        for page in paginator.paginate(Bucket=bucket_name, Prefix=version_prefix):
            for obj in page.get("Contents", []):
                key = obj["Key"]
                filename = key.split("/")[-1]
                if (
                    key.endswith("/")
                    or filename.startswith(".")
                    or any(filename.lower().endswith(e) for e in EXCLUDED_EXTENSIONS)
                ):
                    continue

                path_parts = key.replace(version_prefix, "").split("/")
                lang = path_parts[0] if len(path_parts) > 1 and path_parts[0] in files_by_lang else "root"

                files_by_lang[lang].append({
                    "name": filename,
                    "key": key,
                    "size": obj["Size"],
                    "size_fmt": _fmt_size(obj["Size"]),
                    "last_modified": obj["LastModified"],
                    "icon": _icon(filename),
                    "path": "/".join(path_parts),
                })

        for files in files_by_lang.values():
            files.sort(key=lambda f: f["name"])

    except Exception as exc:
        print(f"Error listing files for version {version}: {exc}")

    return files_by_lang


# ---------------------------------------------------------------------------
# HTML generation via Jinja2
# ---------------------------------------------------------------------------

def build_lang_config(t: Dict[str, str]) -> Dict[str, Dict]:
    """Build the per-language display config injected into the template."""
    return {
        "de":   {"name": t.get("lang_de",   "Deutsch"),         "code": "DE"},
        "fr":   {"name": t.get("lang_fr",   "Français"),        "code": "FR"},
        "it":   {"name": t.get("lang_it",   "Italiano"),        "code": "IT", "status": "draft"},
        "en":   {"name": t.get("lang_en",   "English"),         "code": "EN", "status": "draft"},
        "root": {"name": t.get("lang_root", "Common files"),    "code": "📄"},
    }


def generate_html(
    versions: List[str],
    version_files: Dict[str, Dict],
    bucket_name: str,
    cloudfront_domain: str,
    t: Dict[str, str],
    lang: str,
) -> str:
    base_url = (
        f"https://{cloudfront_domain}"
        if cloudfront_domain
        else f"https://{bucket_name}.s3.amazonaws.com"
    )

    env = Environment(
        loader=FileSystemLoader(str(TEMPLATE_DIR)),
        autoescape=select_autoescape(["html"]),
        trim_blocks=True,
        lstrip_blocks=True,
    )
    template = env.get_template("version_index.html.j2")

    return template.render(
        lang=lang,
        t=t,
        groups=group_versions_by_minor(versions),
        version_files=version_files,
        base_url=base_url,
        lang_config=build_lang_config(t),
        generated_at=datetime.now(timezone.utc).strftime("%d/%m/%Y à %H:%M"),
    )


# ---------------------------------------------------------------------------
# S3 upload
# ---------------------------------------------------------------------------

def upload_to_s3(s3_client, bucket_name: str, html_content: str, key: str):
    try:
        s3_client.put_object(
            Bucket=bucket_name,
            Key=key,
            Body=html_content.encode("utf-8"),
            ContentType="text/html",
            CacheControl="max-age=300",
        )
        print(f"✅ Uploaded s3://{bucket_name}/{key}")
    except Exception as exc:
        print(f"❌ Upload failed: {exc}")
        sys.exit(1)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

SUPPORTED_LANGS = ("de", "fr", "it", "en")


@click.command("process")
@click.option(
    "--lang", default="en", show_default=True,
    type=click.Choice(SUPPORTED_LANGS),
    help="Language for the index page UI.",
)
@click.option(
    "--no-upload", is_flag=True,
    help="Write index.html locally instead of uploading to S3.",
)
def main(lang: str, no_upload: bool):
    """Generate and upload (or save) the versioned documentation index page."""

    bucket_name = os.getenv("S3_BUCKET")
    if not bucket_name:
        print("❌  S3_BUCKET environment variable is not set.")
        sys.exit(2)

    profile_name      = os.getenv("AWS_PROFILE")
    cloudfront_domain = os.getenv("CLOUDFRONT_DOMAIN", "dubious.cloud")
    prefix            = os.getenv("S3_PREFIX", "datamodel/")

    if cloudfront_domain in ("", "none", "None"):
        cloudfront_domain = None
    if os.getenv("CI"):
        profile_name = None

    # --- Load translations ---
    print(f"📖 Loading translations [{lang}] from {TRANSLATION_FILE.relative_to(REPO_ROOT)} …")
    t = load_translations(TRANSLATION_FILE, lang)

    print(f"🔍 Scanning bucket: {bucket_name}  prefix: {prefix}")
    s3_client = get_s3_client(profile_name)

    # --- List versions ---
    print("📋 Listing versions …")
    versions = list_versions(s3_client, bucket_name, prefix)
    print(f"✅ {len(versions)} patch version(s): {', '.join(versions)}")

    groups = group_versions_by_minor(versions)
    print(f"📐 Schema versions (x.y): {', '.join(groups.keys())}")

    # --- Scan files for every patch version ---
    print("📂 Scanning files …")
    version_files: Dict[str, Dict] = {}
    for version in versions:
        files_by_lang = get_files_for_version(s3_client, bucket_name, version, prefix)
        version_files[version] = files_by_lang
        total = sum(len(f) for f in files_by_lang.values())
        summary = ", ".join(f"{l}: {len(f)}" for l, f in files_by_lang.items() if f)
        print(f"  📋 {version}: {total} file(s) ({summary})")

    # --- Render ---
    print("🎨 Rendering template …")
    html_content = generate_html(
        versions, version_files, bucket_name, cloudfront_domain, t, lang
    )

    # --- Output ---
    if not no_upload:
        index_key = f"{prefix}index.html"
        print("⬆️  Uploading …")
        upload_to_s3(s3_client, bucket_name, html_content, index_key)
        target = (
            f"https://{cloudfront_domain}/{index_key}"
            if cloudfront_domain
            else f"https://{bucket_name}.s3.amazonaws.com/{index_key}"
        )
        print(f"🎉 Done → {target}")
    else:
        out = Path("index.html")
        out.write_text(html_content, encoding="utf-8")
        print(f"💾 Written locally to {out}")


if __name__ == "__main__":
    main()
