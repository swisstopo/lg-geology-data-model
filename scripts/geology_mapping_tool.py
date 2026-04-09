#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Geology Mapping Tool Generator
Generates an Excel mapping helper from datamodel.yaml and ESRI exports.
Using the CD domains
"""


import json
import sys
from pathlib import Path

import click
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from rich.console import Console
from rich.progress import BarColumn, Progress, SpinnerColumn, TaskProgressColumn, TextColumn

__version__= '4.0'


console = Console()

# Sheet name overrides (ESRI names → display names, max 31 chars)
NAME_MAPPING = {
    "Unconsolidated_Deposits_PT": "Deposits_PT",
    "Instability_Structures_PT": "Instability_PT",
    "Instability_Structures_L": "Instability_L",
    "Instabilities_PLG": "Instability_PLG",
    "Glacial_Structures_PT": "Glacial_PT",
    "Glacial_and_Periglacial_Structu": "Glacial/Periglacial",
    "Glacial_Structures_PLG": "Glacial_PLG",
    "Erosional_Structures_PT": "Erosion_PT",
    "Erosional_Structures_L": "Erosion_L",
    "Karstic_Structures_PT": "Karst_PT",
    "Karstic_Structures_PLG": "Karst_PLG",
    "Alluvial_and_Lacustrine_Structu": "Alluvial/Lacustrine",
    "Deformation_Structures_PT": "Deformation_PT",
    "Deformation_Structures_L": "Deformation_L",
    "Deformation_Structures_PLG": "Deformation_PLG",
    "Tectonic_Boundaries_L": "Tectonic_L",
    "Folds_PT": "Folds_PT",
    "Lineation_PT": "Lineation_PT",
    "Planar_Structures_PT": "Planar_PT",
    "Anomalies_PT": "Anomaly_PT",
    "Indication_of_Resources_PT": "Resources_PT",
    "Mineralised_Zone_L": "Mineral_L",
    "Sedimentary_Structures_PT": "Sediment_PT",
    "Type_Localities_PT": "Locality_PT",
    "Prominent_Lithological_Features": "Lithology",
    "Miscellaneous_PT": "Misc_PT",
    "Geological_Outlines_L": "Geology_L",
    "Slope_Bedrock_PT": "Slope_PT",
    "Contour_Lines_Bedrock_L": "Contour_Bedrock_L",
    "Modelled_Water_Table_PT": "WaterTable_PT",
    "Contour_Lines_Hydro_L": "Contour_Hydro_L",
    "Archaeology_PT": "Arch_PT",
    "Archaeology_L": "Arch_L",
    "Archaeology_PLG": "Arch_PLG",
    "Exploitation_Geomaterials_PT": "Geomaterials_PT",
    "Exploitation_Geomaterials_L": "Geomaterials_L",
    "Exploitation_Geomaterials_PLG": "Geomaterials_PLG",
    "Boreholes_PT": "Boreholes_PT",
    "Artificial_Surface_Modification": "Surface_Mod",
    "Construction_PT": "Construct_PT",
    "Construction_L": "Construct_L",
    "Palaeohydrology_L": "PalaeoHydro_L",
    "Subsurface_Water_L": "SubWater_L",
    "Surface_Water_PT": "SurfWater_PT",
    "Surface_Water_L": "SurfWater_L",
    "Surface_Water_PLG": "SurfWater_PLG",
}

# (prefix, class_name) pairs — prefix used to filter subtypes by code range
CLASSES_CONFIG = [
    ("12901", "Fossils_PT"),
    ("14401", "Unconsolidated_Deposits_PT"),
    ("11601", "Instability_Structures_PT"),
    ("11701", "Instability_Structures_L"),
    ("15801", "Instabilities_PLG"),
    ("11201", "Glacial_Structures_PT"),
    ("11301", "Glacial_and_Periglacial_Structures_L"),
    ("11401", "Glacial_Structures_PLG"),
    ("11001", "Erosional_Structures_PT"),
    ("11101", "Erosional_Structures_L"),
    ("11901", "Karstic_Structures_PT"),
    ("12001", "Karstic_Structures_PLG"),
    ("10901", "Alluvial_and_Lacustrine_Structures_L"),
    ("14601", "Deformation_Structures_PT"),
    ("14701", "Deformation_Structures_L"),
    ("14801", "Deformation_Structures_PLG"),
    ("14901", "Tectonic_Boundaries_L"),
    ("13601", "Folds_PT"),
    ("13701", "Lineation_PT"),
    ("13801", "Planar_Structures_PT"),
    ("12801", "Anomalies_PT"),
    ("13201", "Indication_of_Resources_PT"),
    ("13301", "Mineralised_Zone_L"),
    ("13401", "Sedimentary_Structures_PT"),
    ("13501", "Type_Localities_PT"),
    ("13101", "Prominent_Lithological_Features_L"),
    ("15501", "Miscellaneous_PT"),
    ("13001", "Geological_Outlines_L"),
    ("14201", "Slope_Bedrock_PT"),
    ("13901", "Contour_Lines_Bedrock_L"),
    ("14101", "Modelled_Water_Table_PT"),
    ("14001", "Contour_Lines_Hydro_L"),
    ("10101", "Archaeology_PT"),
    ("10201", "Archaeology_L"),
    ("10301", "Archaeology_PLG"),
    ("10601", "Exploitation_Geomaterials_PT"),
    ("10701", "Exploitation_Geomaterials_L"),
    ("10801", "Exploitation_Geomaterials_PLG"),
    ("10501", "Boreholes_PT"),
    ("10401", "Artificial_Surface_Modifications_PLG"),
    ("12101", "Construction_PT"),
    ("12201", "Construction_L"),
    ("12301", "Palaeohydrology_L"),
    ("12401", "Subsurface_Water_L"),
    ("12501", "Surface_Water_PT"),
    ("12601", "Surface_Water_L"),
    ("12701", "Surface_Water_PLG"),
]

# Styles
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LOCKED_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
EDITABLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

DATA_ROWS = 1000  # number of data rows per sheet


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_datamodel(datamodel_path: Path) -> tuple[Path, dict]:
    """Parse datamodel.yaml; return (sources_dir, {class_name: [cd_attr_dicts]})."""
    with open(datamodel_path, encoding="utf-8") as f:
        datamodel = yaml.safe_load(f)

    model = datamodel.get("model", {})
    sources_dir = Path(model["sources_dir"])

    class_cd_attributes: dict[str, list] = {}
    for theme in datamodel.get("themes", []):
        for cls in theme.get("classes", []):
            class_name = cls.get("name")
            cd_attrs = [
                {
                    "name": attr["name"],
                    "cd_domain": attr.get("value", ""),
                    "desc_de": attr.get("description", {}).get("de", ""),
                    "desc_fr": attr.get("description", {}).get("fr", ""),
                    "mandatory": attr.get("mandatory", False),
                }
                for attr in cls.get("attributes", [])
                if attr.get("att_type") == "CD"
                and attr.get("value") not in (None, "N/A", "dummy", "DUMMY")
            ]
            if cd_attrs:
                class_cd_attributes[class_name] = cd_attrs

    return sources_dir, class_cd_attributes


def load_json(sources_dir: Path, filename: str, *, is_list: bool = False):
    """Load a JSON file from the sources directory; return {} or [] on failure."""
    filepath = sources_dir / filename
    try:
        with open(filepath, encoding="utf-8") as f:
            data = json.load(f)
        count = len(data) if isinstance(data, (dict, list)) else "OK"
        console.print(f"   [green]✓[/] {filename}: {count}")
        return data
    except FileNotFoundError:
        console.print(f"   [yellow]⚠[/]  {filename}: not found")
        return [] if is_list else {}
    except Exception as e:
        console.print(f"   [red]✗[/]  {filename}: {e}")
        return [] if is_list else {}


def load_gcoverp(sources_dir: Path) -> tuple[dict, dict]:
    """Load coded domains and subtypes from gcoverp_export_simple.json."""
    gcoverp_file = sources_dir / "gcoverp_export_simple.json"
    coded_domains: dict = {}
    subtypes: dict = {}

    try:
        with open(gcoverp_file, encoding="utf-8") as f:
            gcoverp_data = json.load(f)

        subtypes = gcoverp_data.get("subtypes", {})
        console.print(f"   [green]✓[/] subtypes: {len(subtypes)} entries")

        for domain_name, domain_data in gcoverp_data.get("coded_domain", {}).items():
            if isinstance(domain_data, dict) and "codedValues" in domain_data:
                coded_domains[domain_name] = domain_data["codedValues"]
        console.print(f"   [green]✓[/] coded domains: {len(coded_domains)} domains")

    except FileNotFoundError:
        console.print(f"   [yellow]⚠[/]  gcoverp_export_simple.json: not found")
    except Exception as e:
        console.print(f"   [red]✗[/]  gcoverp_export_simple.json: {e}")

    return coded_domains, subtypes


# ---------------------------------------------------------------------------
# Sheet helpers
# ---------------------------------------------------------------------------

def _set_header(ws, col: int, text: str) -> None:
    cell = ws.cell(1, col, text)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT


def create_reference_sheet(wb: Workbook, sheet_name: str, data_dict: dict) -> int:
    """Create a Ref_* sheet with Code / Label / Display columns. Returns last data row."""
    ws = wb.create_sheet(sheet_name)
    for col, header in enumerate(["Code", "Label", "Display"], start=1):
        _set_header(ws, col, header)

    # Push codes < 1_000_000 to the end; sort the rest alphabetically by label
    items = sorted(data_dict.items(), key=lambda x: (int(x[0]) < 1_000_000, str(x[1])))

    for row, (code, label) in enumerate(items, start=2):
        ws.cell(row, 1, int(code))
        ws.cell(row, 2, str(label))
        ws.cell(row, 3, f"{code}: {label}")

    return row  # last written row


def add_dropdown(ws, col_range: str, ref_sheet: str, ref_range: str,
                 prompt: str, title: str) -> None:
    dv = DataValidation(type="list", formula1=f"={ref_sheet}!${ref_range}", allow_blank=True)
    dv.prompt = prompt
    dv.promptTitle = title
    ws.add_data_validation(dv)
    dv.add(col_range)


def extract_subtypes_by_prefix(subtypes_dict: dict, prefix: str) -> dict:
    return {k: v for k, v in subtypes_dict.items() if str(k).startswith(str(prefix))}


def _col(n: int) -> str:
    """Return the Excel column letter for 1-based column index (handles AA, AB, …)."""
    return get_column_letter(n)


def _extract_code_formula(display_col: str, row: int) -> str:
    return (
        f'=IF({display_col}{row}="",'
        f'"",VALUE(LEFT({display_col}{row},FIND(":",{display_col}{row})-1)))'
    )


# ---------------------------------------------------------------------------
# Main sheet builders
# ---------------------------------------------------------------------------

def build_bedrock_sheet(wb: Workbook, gmu_last_row: int, tecto_last_row: int) -> None:
    ws = wb.create_sheet("Bedrock_PLG", 0)
    headers = [
        "User_Code", "GEOL_MAPPING_UNIT_Display", "GEOL_MAPPING_UNIT",
        "TECTO_Display", "TECTO",
        "LITSTRAT_FORMATION_BANK_Display",
        "CHRONO_TOP_Display", "CHRONO_BASE_Display",
        "LITHO_MAIN_Display", "LITHO_SEC_Display", "LITHO_TER_Display",
    ]
    for i, h in enumerate(headers, 1):
        _set_header(ws, i, h)

    for i, w in enumerate([15, 50, 20, 40, 15, 50, 35, 35, 35, 35, 35], 1):
        ws.column_dimensions[_col(i)].width = w
    for col in ["C", "E"]:
        ws.column_dimensions[col].hidden = True

    add_dropdown(ws, f"B2:B{DATA_ROWS}", "Ref_GeolMappingUnit",
                 f"C$2:C${gmu_last_row}", "Sélectionnez une unité", "GEOL_MAPPING_UNIT")
    add_dropdown(ws, f"D2:D{DATA_ROWS}", "Ref_Tecto",
                 f"C$2:C${tecto_last_row}", "Sélectionnez une unité tectonique", "TECTO")

    lookup_cols = [
        ("F", "B", "Ref_LithoStratFormation"),
        ("G", "C", "Ref_Chrono"),
        ("H", "D", "Ref_Chrono"),
        ("I", "E", "Ref_Litho"),
        ("J", "F", "Ref_Litho"),
        ("K", "G", "Ref_Litho"),
    ]
    for row in range(2, DATA_ROWS + 1):
        ws[f"A{row}"].fill = EDITABLE_FILL
        ws[f"B{row}"].fill = EDITABLE_FILL
        ws[f"C{row}"] = _extract_code_formula("B", row)
        ws[f"D{row}"].fill = EDITABLE_FILL
        ws[f"E{row}"] = _extract_code_formula("D", row)
        for target_col, source_col, ref_sheet in lookup_cols:
            ws[f"{target_col}{row}"] = (
                f'=IF(C{row}="","",IFERROR('
                f'INDEX({ref_sheet}!$C:$C,MATCH(VLOOKUP(C{row},Ref_GMU_Attributes!$A:${source_col},'
                f'COLUMN({source_col}:${source_col}),FALSE),{ref_sheet}!$A:$A,0)),""))'
            )
            ws[f"{target_col}{row}"].fill = LOCKED_FILL

    for i, code in enumerate(["BEDROCK_001", "BEDROCK_002"], 2):
        ws[f"A{i}"] = code


def build_unco_sheet(wb: Workbook, litstrat_unco_last_row: int,
                     litho_last_row: int, chrono_last_row: int) -> None:
    ws = wb.create_sheet("Unconsolidated_Deposits_PLG", 1)
    headers = [
        "User_Code",
        "LITSTRAT_Display", "LITSTRAT",
        "LITHO_Display", "LITHO",
        "CHRONO_TOP_Display", "CHRONO_TOP",
        "CHRONO_BASE_Display", "CHRONO_BASE",
    ]
    for i, h in enumerate(headers, 1):
        _set_header(ws, i, h)
    for col in ["C", "E", "G", "I"]:
        ws.column_dimensions[col].hidden = True
    for i in range(1, 10):
        ws.column_dimensions[_col(i)].width = 15 if i in (1, 3, 5, 7, 9) else 50

    add_dropdown(ws, f"B2:B{DATA_ROWS}", "Ref_LithoStratUnco",
                 f"C$2:C${litstrat_unco_last_row}", "Sélectionnez une unité", "LITSTRAT")
    add_dropdown(ws, f"D2:D{DATA_ROWS}", "Ref_Litho",
                 f"C$2:C${litho_last_row}", "Sélectionnez une lithologie", "LITHO")
    add_dropdown(ws, f"F2:F{DATA_ROWS}", "Ref_Chrono",
                 f"C$2:C${chrono_last_row}", "Chronostratigraphie (top)", "CHRONO_TOP")
    add_dropdown(ws, f"H2:H{DATA_ROWS}", "Ref_Chrono",
                 f"C$2:C${chrono_last_row}", "Chronostratigraphie (base)", "CHRONO_BASE")

    for row in range(2, DATA_ROWS + 1):
        for col in ["A", "B", "D", "F", "H"]:
            ws[f"{col}{row}"].fill = EDITABLE_FILL
        ws[f"C{row}"] = _extract_code_formula("B", row)
        ws[f"E{row}"] = _extract_code_formula("D", row)
        ws[f"G{row}"] = _extract_code_formula("F", row)
        ws[f"I{row}"] = _extract_code_formula("H", row)

    for i, code in enumerate(["UNCO_001", "UNCO_002"], 2):
        ws[f"A{i}"] = code


def build_class_sheet(
    wb: Workbook,
    sheet_idx: int,
    class_name: str,
    cd_attrs: list,
    subtype_refs: dict,
    cd_refs: dict,
) -> None:
    sheet_name = NAME_MAPPING.get(class_name, class_name[:31])
    ws = wb.create_sheet(sheet_name, sheet_idx)

    # Build column list: User_Code | KIND_Display | KIND | [ATTR_Display | ATTR]… | Notes
    headers = ["User_Code", "KIND_Display", "KIND"]
    for cd_attr in cd_attrs:
        attr_name = cd_attr["name"].upper()
        headers += [f"{attr_name}_Display", attr_name]
    headers.append("Notes")

    for i, h in enumerate(headers, 1):
        _set_header(ws, i, h)

    # Column widths — hide code columns, widen display/user/notes
    for i, h in enumerate(headers, 1):
        col_letter = _col(i)
        if h in ("Notes", "User_Code") or h.endswith("_Display"):
            ws.column_dimensions[col_letter].width = 20 if i <= 3 else 35
        else:
            ws.column_dimensions[col_letter].hidden = True

    # KIND dropdown
    if class_name in subtype_refs:
        ref_name = f"Ref_KIND_{class_name}"[:31]
        add_dropdown(
            ws, f"B2:B{DATA_ROWS}", ref_name,
            f"C$2:C${subtype_refs[class_name]}",
            "Sélectionnez un KIND", "KIND",
        )

    # CD attribute dropdowns (display columns are even-indexed starting at 4)
    for attr_idx, cd_attr in enumerate(cd_attrs):
        col_idx = 4 + attr_idx * 2  # 1-based index of the _Display column
        cd_domain = cd_attr["cd_domain"]
        attr_name = cd_attr["name"].upper()
        if cd_domain in cd_refs:
            ref_info = cd_refs[cd_domain]
            col_letter = _col(col_idx)
            add_dropdown(
                ws, f"{col_letter}2:{col_letter}{DATA_ROWS}",
                ref_info["sheet_name"],
                f"C$2:C${ref_info['last_row']}",
                cd_attr["desc_de"], attr_name,
            )

    # Row formulas and fills
    notes_col = _col(len(headers))
    for row in range(2, DATA_ROWS + 1):
        ws[f"A{row}"].fill = EDITABLE_FILL
        ws[f"B{row}"].fill = EDITABLE_FILL
        ws[f"C{row}"] = _extract_code_formula("B", row)

        for attr_idx in range(len(cd_attrs)):
            col_idx = 4 + attr_idx * 2
            disp_col = _col(col_idx)
            code_col = _col(col_idx + 1)
            ws[f"{disp_col}{row}"].fill = EDITABLE_FILL
            ws[f"{code_col}{row}"] = _extract_code_formula(disp_col, row)

        ws[f"{notes_col}{row}"].fill = EDITABLE_FILL

    ws["A2"] = f"{class_name.upper().replace('_', '')[:10]}_001"


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

@click.command()
@click.version_option(__version__)
@click.argument("datamodel", default="datamodel.yaml",
                type=click.Path(exists=True, dir_okay=False, path_type=Path))
@click.option("--output", "-o", default="outputs/geology_mapping_tool.xlsx",
              type=click.Path(path_type=Path),
              help="Output Excel file  [default: outputs/geology_mapping_tool.xlsx]",
              show_default=False)
@click.option("--sources-dir", "-s", default=None,
              type=click.Path(exists=True, file_okay=False, path_type=Path),
              help="Override sources directory from datamodel.yaml")
def main(datamodel: Path, output: Path, sources_dir: Path | None) -> None:
    """Generate the GeoCover Excel mapping tool from DATAMODEL (default: datamodel.yaml)."""
    console.rule("[bold blue]Geology Mapping Tool Generator v4.0")

    # --- Load datamodel ---
    console.print("\n[bold]Loading datamodel.yaml…")
    try:
        dm_sources_dir, class_cd_attributes = load_datamodel(datamodel)
    except (KeyError, FileNotFoundError) as e:
        console.print(f"[red]✗[/] Cannot load {datamodel}: {e}")
        sys.exit(1)

    sources = sources_dir or dm_sources_dir
    if not sources.is_dir():
        console.print(f"[red]✗[/] Sources directory not found: {sources}")
        sys.exit(2)

    console.print(
        f"   [green]✓[/] {len(class_cd_attributes)} classes with CD attributes\n"
        f"   [green]✓[/] sources: {sources}"
    )

    # --- Load JSON data files ---
    console.print("\n[bold]Loading data files…")
    geol_mapping_units     = load_json(sources, "Geol_Mapping_Unit.json")
    geol_mapping_unit_att  = load_json(sources, "Geol_Mapping_Unit_Att.json", is_list=True)
    litho                  = load_json(sources, "Litho.json")
    chrono                 = load_json(sources, "Chrono.json")
    tecto                  = load_json(sources, "Tecto.json")
    litstrat_fm            = load_json(sources, "Litstrat_Formation_Bank.json")
    litstrat_unco          = load_json(sources, "Litstrat_Unco.json")
    system_data            = load_json(sources, "System.json")

    console.print("\n[bold]Loading coded domains and subtypes…")
    coded_domains, subtypes = load_gcoverp(sources)

    # --- Build workbook ---
    wb = Workbook()

    # Reference sheets
    console.print("\n[bold]Creating reference sheets…")
    gmu_last_row        = create_reference_sheet(wb, "Ref_GeolMappingUnit", geol_mapping_units)
    litho_last_row      = create_reference_sheet(wb, "Ref_Litho", litho)
    chrono_last_row     = create_reference_sheet(wb, "Ref_Chrono", chrono)
    tecto_last_row      = create_reference_sheet(wb, "Ref_Tecto", tecto)
    litstrat_fm_last_row  = create_reference_sheet(wb, "Ref_LithoStratFormation", litstrat_fm)
    litstrat_unco_last_row = create_reference_sheet(wb, "Ref_LithoStratUnco", litstrat_unco)
    create_reference_sheet(wb, "Ref_System", system_data)

    # GMU attributes sheet
    ws_ref_gmu_att = wb.create_sheet("Ref_GMU_Attributes")
    gmu_att_headers = [
        "GEOL_MAPPING_UNIT", "LITSTRAT_FORMATION_BANK",
        "CHRONO_TOP", "CHRONO_BASE", "LITHO_MAIN", "LITHO_SEC", "LITHO_TER",
    ]
    for i, h in enumerate(gmu_att_headers, 1):
        _set_header(ws_ref_gmu_att, i, h)
    for row, att in enumerate(geol_mapping_unit_att, start=2):
        ws_ref_gmu_att.cell(row, 1, att["GEOL_MAPPING_UNIT"])
        ws_ref_gmu_att.cell(row, 2, att["LITSTRAT_FORMATION_BANK"])
        ws_ref_gmu_att.cell(row, 3, att["CHRONO_TOP"])
        ws_ref_gmu_att.cell(row, 4, att["CHRONO_BASE"])
        ws_ref_gmu_att.cell(row, 5, att["LITHO_MAIN"])
        ws_ref_gmu_att.cell(row, 6, att["LITHO_SEC"] if att["LITHO_SEC"] != 0 else "")
        ws_ref_gmu_att.cell(row, 7, att["LITHO_TER"] if att["LITHO_TER"] != 0 else "")

    # Coded domain reference sheets
    console.print("\n[bold]Creating coded domain sheets…")
    cd_refs: dict = {}
    for cd_name, cd_values in coded_domains.items():
        if not cd_values:
            continue
        ref_name = f"Ref_CD_{cd_name}"[:31]
        try:
            last_row = create_reference_sheet(wb, ref_name, cd_values)
            cd_refs[cd_name] = {"sheet_name": ref_name, "last_row": last_row}
        except Exception as e:
            console.print(f"   [yellow]⚠[/]  Cannot create {ref_name}: {e}")
    console.print(f"   [green]✓[/] {len(cd_refs)} coded domain sheets")

    # KIND (subtype) reference sheets
    console.print("\n[bold]Extracting subtypes by class…")
    subtype_refs: dict = {}
    for prefix, class_name in CLASSES_CONFIG:
        kind_values = extract_subtypes_by_prefix(subtypes, prefix)
        if kind_values:
            ref_name = f"Ref_KIND_{class_name}"[:31]
            subtype_refs[class_name] = create_reference_sheet(wb, ref_name, kind_values)
    console.print(f"   [green]✓[/] {len(subtype_refs)} KIND sheets")

    # Main sheets: Bedrock + Unconsolidated
    console.print("\n[bold]Building main sheets…")
    build_bedrock_sheet(wb, gmu_last_row, tecto_last_row)
    build_unco_sheet(wb, litstrat_unco_last_row, litho_last_row, chrono_last_row)
    console.print("   [green]✓[/] Bedrock_PLG, Unconsolidated_Deposits_PLG")

    # Per-class sheets
    console.print("\n[bold]Building class sheets…")
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Classes…", total=len(CLASSES_CONFIG))
        for sheet_idx, (prefix, class_name) in enumerate(CLASSES_CONFIG, start=2):
            cd_attrs = class_cd_attributes.get(class_name, [])
            build_class_sheet(wb, sheet_idx, class_name, cd_attrs, subtype_refs, cd_refs)
            progress.advance(task)

    # Clean up default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    console.rule("[bold green]Done")
    console.print(f"\n[bold]Output:[/] {output}")
    console.print(f"\n[bold]Summary:[/]")
    console.print(f"  2 main sheets  (Bedrock, Unconsolidated)")
    console.print(f"  {len(CLASSES_CONFIG)} class sheets")
    console.print(f"  {len(subtype_refs)} KIND reference sheets")
    console.print(f"  {len(cd_refs)} coded domain reference sheets")
    console.print(f"  {len(wb.sheetnames)} sheets total")


if __name__ == "__main__":
    main()
